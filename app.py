import streamlit as st
import os
import sys
from datetime import date, datetime
from PIL import Image
import io

sys.path.insert(0, os.path.dirname(__file__))

from database import (
    init_db, get_connection,
    get_sectores, save_sector, update_sector, delete_sector,
    get_materiales_catalogo, save_material_catalogo, update_material_catalogo, delete_material_catalogo,
    get_subcontratas, save_subcontrata, update_subcontrata, delete_subcontrata,
    get_partidas, save_partida, update_partida, delete_partida,
    get_registro_diario, get_or_create_registro_diario,
    save_avance, get_avances, get_avance, update_avance, delete_avance,
    get_registros_fechas,
    save_field_note, get_field_notes, update_field_note, delete_field_note,
    save_photo, get_photos, update_photo, delete_photo,
    save_material, get_materials, update_material, delete_material,
    get_all_fechas,
    get_config, set_config, get_all_config,
    get_avances_por_rango, get_registros_por_rango,
    get_materials_por_rango, get_field_notes_por_rango,
)
from export import generar_informe, generar_informe_semanal

st.set_page_config(
    page_title="CONTROL DE OBRA",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="collapsed",
)

init_db()

CLIMA_OPTS = ["Soleado", "Nublado", "Lluvia", "Lluvia Intensa"]

PHOTOS_DIR = os.path.join(os.path.dirname(__file__), "photos")
os.makedirs(PHOTOS_DIR, exist_ok=True)

st.markdown(
    """
<style>
    .stApp { background-color: #101010; }
    .block-container { padding-top: 1rem; padding-bottom: 4rem; }
    h1, h2, h3 { color: #FFCD00 !important; font-weight: 800 !important; font-family: 'Arial Black', sans-serif !important; }
    body, p, label, span, div, .stMarkdown, .stText, .st-bm, .st-bn, .st-bo {
        color: #FFFFFF !important;
    }
    .stButton > button {
        width: 100%;
        height: 2.8em;
        font-size: 1rem;
        font-weight: 900 !important;
        border-radius: 8px;
        background-color: #FFCD00;
        color: #000000 !important;
        border: none;
        box-shadow: 0 2px 6px rgba(255,205,0,0.3);
        text-shadow: 1px 1px 2px rgba(0,0,0,0.3);
    }
    .stButton > button:hover {
        background-color: #E6B800;
        color: #000000 !important;
        border: none;
        box-shadow: 0 4px 12px rgba(255,205,0,0.5);
        text-shadow: 1px 1px 2px rgba(0,0,0,0.3);
    }
    .stButton > button:active {
        background-color: #CCA300;
        color: #000000 !important;
    }
    .card {
        background: #333333;
        padding: 1rem;
        border-radius: 10px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.3);
        margin-bottom: 0.8rem;
        color: #FFFFFF !important;
        border-left: 4px solid #FFCD00;
    }
    .card strong {
        color: #FFCD00 !important;
    }
    .card-sm {
        background: #333333;
        padding: 0.6rem 1rem;
        border-radius: 8px;
        box-shadow: 0 1px 4px rgba(0,0,0,0.2);
        margin-bottom: 0.5rem;
        color: #FFFFFF !important;
    }
    .stat-box {
        background: #222222;
        padding: 0.8rem;
        border-radius: 8px;
        text-align: center;
        margin: 0.3rem;
        color: #FFCD00 !important;
        border: 1px solid #FFCD00;
    }
    .stat-box strong {
        color: #FFFFFF !important;
    }
    .admin-btn > button {
        background-color: #444444;
        font-size: 0.85rem;
        height: 2.2em;
        color: #FFFFFF !important;
        box-shadow: none;
        text-shadow: none;
    }
    .admin-btn > button:hover {
        background-color: #555555;
        color: #FFCD00 !important;
    }
    .del-btn > button {
        background-color: #8B0000;
        color: #FFFFFF !important;
        box-shadow: none;
        text-shadow: none;
    }
    .del-btn > button:hover {
        background-color: #A52A2A;
        color: #FFFFFF !important;
    }
    div[data-testid="stHorizontalBlock"]:last-of-type {
        position: sticky !important;
        bottom: 0 !important;
        background: #1a1a1a !important;
        padding: 0.5rem 0 !important;
        z-index: 999 !important;
        box-shadow: 0 -2px 8px rgba(0,0,0,0.3);
        border-top: 2px solid #FFCD00;
    }
    hr {
        border-color: #FFCD00 !important;
        border-width: 2px !important;
    }
    .stAlert {
        background-color: #333333 !important;
        color: #FFCD00 !important;
        border-left: 4px solid #FFCD00 !important;
    }
    .stSuccess {
        background-color: #222222 !important;
        color: #FFCD00 !important;
        border-left: 4px solid #FFCD00 !important;
    }
    .stWarning {
        background-color: #333333 !important;
        color: #FFCD00 !important;
    }
    .stError {
        background-color: #333333 !important;
        color: #FF4444 !important;
    }
    .stSelectbox, .stTextInput, .stNumberInput, .stDateInput, .stTextArea {
        color: #FFFFFF !important;
        background-color: #000000 !important;
    }
    .st-dx {
        background-color: #FFCD00 !important;
    }
    .stSlider > div > div {
        color: #FFCD00 !important;
    }
    .st-emotion-cache-1v0mbdj {
        color: #FFCD00 !important;
    }
    .st-emotion-cache-15hul6a {
        background-color: #000000 !important;
        border: 1px solid #555555;
        color: #FFFFFF !important;
    }
    .st-emotion-cache-1n76uvr {
        color: #FFFFFF !important;
    }
    .st-emotion-cache-1h9ksn3 {
        color: #FFCD00 !important;
    }
    [data-testid="stExpander"] {
        background-color: #222222;
        border: 1px solid #444444;
        border-radius: 8px;
    }
    [data-testid="stExpander"] summary {
        color: #FFCD00;
        font-weight: 700;
    }
    [data-testid="stTable"] {
        color: #FFFFFF !important;
    }
    .stTabs [data-baseweb="tab-list"] {
        background-color: #222222;
        border-radius: 8px;
        padding: 4px;
    }
    .stTabs [data-baseweb="tab"] {
        color: #FFFFFF !important;
        font-weight: 600;
    }
    .stTabs [aria-selected="true"] {
        color: #000000 !important;
        background-color: #FFCD00 !important;
        border-radius: 6px;
    }
    [data-testid="stForm"] {
        background-color: #1a1a1a;
        padding: 1rem;
        border-radius: 10px;
        border: 1px solid #333333;
    }
    .stSubheader {
        color: #FFCD00 !important;
    }
    .stMarkdown strong {
        color: #FFCD00;
    }
    .st-bb {
        background-color: #FFCD00 !important;
    }
    .st-cb {
        color: #FFCD00 !important;
    }
    input, textarea, select {
        color: #FFFFFF !important;
        background-color: #000000 !important;
    }
    .st-bw {
        background-color: #000000 !important;
    }
    .st-cr {
        color: #FFFFFF !important;
    }
    .st-cn {
        color: #FFFFFF !important;
    }
    .st-dg {
        color: #FFFFFF !important;
    }
    .st-dh {
        color: #FFFFFF !important;
    }
    .st-di {
        color: #FFFFFF !important;
    }
    [data-testid="stFileUploader"] button {
        background-color: #FFCD00 !important;
        color: #000000 !important;
        font-weight: 700 !important;
        border: none !important;
    }
    [data-testid="stFileUploader"] button:hover {
        background-color: #E6B800 !important;
        color: #000000 !important;
    }
    [data-testid="stFileUploader"] button:active {
        background-color: #CCA300 !important;
        color: #000000 !important;
    }
    button[kind="secondary"] {
        background-color: #FFCD00 !important;
        color: #000000 !important;
        font-weight: 700 !important;
        border: none !important;
    }
    button[kind="secondary"]:hover {
        background-color: #E6B800 !important;
        color: #000000 !important;
    }
    button[kind="secondaryFormSubmit"] {
        background-color: #FFCD00 !important;
        color: #000000 !important;
        font-weight: 700 !important;
    }
    button[kind="secondaryFormSubmit"]:hover {
        background-color: #E6B800 !important;
        color: #000000 !important;
    }
</style>
""",
    unsafe_allow_html=True,
)

if "page" not in st.session_state:
    st.session_state.page = "inicio"
if "current_fecha" not in st.session_state:
    st.session_state.current_fecha = date.today().isoformat()


def nav_button(label, page, col, key=None):
    with col:
        if st.button(label, key=key, use_container_width=True):
            st.session_state.page = page
            st.rerun()


def show_header():
    st.markdown(
        '<h1 style="text-align:center; margin-bottom:0; color:#FFCD00; font-weight:900; font-size:2.2rem; text-transform:uppercase; letter-spacing:2px;">CONTROL DE OBRA</h1>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<hr style="border:2px solid #FFCD00; border-radius:2px;">',
        unsafe_allow_html=True,
    )


def card(text, class_name="card"):
    st.markdown(f'<div class="{class_name}">{text}</div>', unsafe_allow_html=True)


def back_button():
    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        if st.button("⬅️ Volver al Inicio", key=f"back_{st.session_state.page}", use_container_width=True):
            st.session_state.page = "inicio"
            st.rerun()


def edit_del_buttons(edit_key, del_key, edit_label="✏️", del_label="🗑️"):
    cols = st.columns([1, 1, 8])
    with cols[0]:
        if st.button(edit_label, key=edit_key, use_container_width=True):
            return "edit"
    with cols[1]:
        if st.button(del_label, key=del_key, use_container_width=True):
            return "delete"
    return None


def confirm_dialog(key_prefix, rid, on_confirm):
    st.warning("¿Eliminar este registro?")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ Sí, eliminar", key=f"confirm_{key_prefix}_{rid}"):
            on_confirm(rid)
            st.success("Eliminado")
            for k in list(st.session_state.keys()):
                if k.startswith(key_prefix):
                    del st.session_state[k]
            st.rerun()
    with c2:
        if st.button("❌ Cancelar", key=f"cancel_{key_prefix}_{rid}"):
            for k in list(st.session_state.keys()):
                if k.startswith(key_prefix):
                    del st.session_state[k]
            st.rerun()


# ═══════════════════════════════════════════════
# PAGE: INICIO
# ═══════════════════════════════════════════════
def page_inicio():
    show_header()
    hoy = date.today().isoformat()

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        card(f"<strong>🗓️ Hoy</strong><br>{hoy}", "stat-box")
    with col2:
        partidas = get_partidas()
        card(f"<strong>📋 Partidas</strong><br>{len(partidas)} registradas", "stat-box")
    with col3:
        subc = get_subcontratas()
        card(f"<strong>👥 Subcontratas</strong><br>{len(subc)} registradas", "stat-box")
    with col4:
        reg = get_registro_diario(hoy)
        card(f"<strong>✅ Hoy</strong><br>{'Registrado' if reg else 'Pendiente'}", "stat-box")

    st.markdown("<br>", unsafe_allow_html=True)

    cols = st.columns(3)
    nav_button("📋 Registro Diario", "registro", cols[0])
    nav_button("📦 Materiales", "materiales", cols[1])
    nav_button("📝 Notas de Campo", "notas", cols[2])
    cols = st.columns(3)
    nav_button("📸 Registro Fotográfico", "fotos", cols[0])
    nav_button("📄 Generar Informe", "informe", cols[1])
    nav_button("📅 Informe Semanal", "informe_semanal", cols[2])
    cols = st.columns(3)
    nav_button("⚙️ Administrar", "admin", cols[0])
    nav_button("📈 Avance General", "avance_gral", cols[1])
    nav_button("📊 Ver Historial", "historial", cols[2])

    reg = get_registro_diario(hoy)
    if reg:
        avances = get_avances(reg["id"])
        st.markdown("---")
        st.markdown("#### Último registro de hoy")
        card(
            f"<strong>Clima:</strong> {reg['clima']} | "
            f"<strong>Partidas trabajadas:</strong> {len(avances)}"
        )


# ═══════════════════════════════════════════════
# PAGE: REGISTRO DIARIO (NEW)
# ═══════════════════════════════════════════════
def page_registro():
    show_header()
    st.markdown("## 📋 Registro Diario de Obra")

    fecha = st.date_input("Fecha", value=date.today(), key="reg_fecha")
    fecha_str = fecha.isoformat()
    st.session_state.current_fecha = fecha_str

    rid = get_or_create_registro_diario(fecha_str)
    registro = get_registro_diario(fecha_str)
    sectores = get_sectores()
    sec_opts = {s["nombre"]: s["id"] for s in sectores}
    subcontratas = get_subcontratas()
    subc_opts = {s["nombre"]: s["id"] for s in subcontratas}
    subc_opts["(De la empresa)"] = None

    col1, col2 = st.columns(2)
    with col1:
        clima = st.select_slider("Clima", CLIMA_OPTS, value=registro["clima"] if registro else "Soleado", key="reg_clima")
    with col2:
        obs = st.text_input("Observaciones", value=registro.get("observaciones", "") if registro else "", key="reg_obs")

    if st.button("💾 Guardar datos del día", key="save_reg_dia"):
        conn = get_connection()
        conn.execute("UPDATE registro_diario SET clima=?, observaciones=? WHERE id=?", (clima, obs, rid))
        conn.commit()
        conn.close()
        st.success("Datos del día guardados")
        st.rerun()

    st.divider()

    # ── Add avance ──
    st.markdown("### ➕ Agregar avance de partida")
    partidas = get_partidas()

    if not partidas:
        st.warning("No hay partidas registradas. Ve a ⚙️ Administrar para agregarlas.")
    else:
        with st.form("form_avance"):
            col1, col2 = st.columns(2)
            with col1:
                partida_opts = {f"{p['codigo'] + ' - ' if p['codigo'] else ''}{p['nombre']}": p["id"] for p in partidas}
                p_sel = st.selectbox("Partida", options=list(partida_opts.keys()), key="av_partida")
                pid = partida_opts[p_sel]
            with col2:
                sector_sel = st.selectbox("Sector", options=list(sec_opts.keys()) if sec_opts else ["(Sin sector)"], key="av_sec")
                sec_id = sec_opts[sector_sel] if sec_opts else None

            col1, col2 = st.columns(2)
            with col1:
                subc_sel = st.selectbox("Ejecutado por", options=list(subc_opts.keys()), key="av_subc")
                subc_id = subc_opts[subc_sel]

            col1, col2, col3 = st.columns(3)
            with col1:
                operarios = st.number_input("Operarios", min_value=0, value=0, step=1, key="av_op")
            with col2:
                oficiales = st.number_input("Oficiales", min_value=0, value=0, step=1, key="av_of")
            with col3:
                peones = st.number_input("Peones", min_value=0, value=0, step=1, key="av_pn")

            col1, col2 = st.columns(2)
            with col1:
                horas = st.number_input("Horas trabajadas", min_value=0.0, value=0.0, step=0.5, key="av_hr")
            with col2:
                cant_ejec = st.number_input("Cantidad ejecutada", min_value=0.0, value=0.0, step=0.1, key="av_ce")

            obs_avance = st.text_input("Observaciones", placeholder="Ej: Avance normal, sin contratiempos", key="av_obs")

            if st.form_submit_button("💾 Agregar avance"):
                save_avance(rid, pid, subc_id, sec_id, operarios, oficiales, peones, horas, cant_ejec, obs_avance)
                st.success("Avance agregado")
                st.rerun()

    st.divider()

    # ── List avances ──
    st.markdown(f"### 📋 Avances del día ({fecha_str})")
    avances = get_avances(rid)

    if avances:
        for a in avances:
            aid = a["id"]
            total_personal = a["num_operarios"] + a["num_oficiales"] + a["num_peones"]
            with st.container():
                card(
                    f"<strong>{a['partida_codigo'] + ' - ' if a['partida_codigo'] else ''}{a['partida_nombre']}</strong>"
                    f"<br>🏗️ {a['sector_nombre'] or 'Sin sector'}"
                    f" | 🏢 {a['subcontrata_nombre'] or 'De la empresa'}"
                    f"<br>👥 {total_personal} pers "
                    f"(Op: {a['num_operarios']} / Of: {a['num_oficiales']} / Pn: {a['num_peones']}) "
                    f"| ⏱ {a['horas_trabajadas']}h "
                    f"| 📊 Ejecutado: {a['cantidad_ejecutada']} {a['partida_unidad'] or ''}"
                )
                cols = st.columns([1, 1, 8])
                with cols[0]:
                    if st.button("✏️", key=f"edit_av_{aid}", use_container_width=True):
                        st.session_state.edit_av = aid
                        st.rerun()
                with cols[1]:
                    if st.button("🗑️", key=f"del_av_{aid}", use_container_width=True):
                        st.session_state.del_av = aid
                        st.rerun()

                if st.session_state.get("edit_av") == aid:
                    with st.form(key=f"form_edit_av_{aid}"):
                        col1, col2 = st.columns(2)
                        with col1:
                            subc_opts_edit = {s["nombre"]: s["id"] for s in subcontratas}
                            subc_opts_edit["(De la empresa)"] = None
                            cur_subc = a["subcontrata_nombre"] or "(De la empresa)"
                            s_sel = st.selectbox("Ejecutado por", options=list(subc_opts_edit.keys()), index=list(subc_opts_edit.keys()).index(cur_subc) if cur_subc in subc_opts_edit else 0, key=f"e_subc_{aid}")
                        with col2:
                            cur_sec = a["sector_nombre"] or list(sec_opts.keys())[0] if sec_opts else "(Sin sector)"
                            sec_sel = st.selectbox("Sector", options=list(sec_opts.keys()) if sec_opts else ["(Sin sector)"], index=list(sec_opts.keys()).index(cur_sec) if cur_sec in sec_opts else 0, key=f"e_sec_{aid}")
                        c1, c2, c3 = st.columns(3)
                        with c1:
                            op = st.number_input("Operarios", min_value=0, value=a["num_operarios"], step=1, key=f"e_op_{aid}")
                        with c2:
                            of = st.number_input("Oficiales", min_value=0, value=a["num_oficiales"], step=1, key=f"e_of_{aid}")
                        with c3:
                            pn = st.number_input("Peones", min_value=0, value=a["num_peones"], step=1, key=f"e_pn_{aid}")
                        c1, c2 = st.columns(2)
                        with c1:
                            hr = st.number_input("Horas", min_value=0.0, value=float(a["horas_trabajadas"]), step=0.5, key=f"e_hr_{aid}")
                        with c2:
                            ce = st.number_input("Cant. ejecutada", min_value=0.0, value=float(a["cantidad_ejecutada"]), step=0.1, key=f"e_ce_{aid}")
                        obs_av = st.text_input("Observaciones", value=a["observaciones"] or "", key=f"e_obs_{aid}")
                        if st.form_submit_button("💾 Guardar"):
                            update_avance(aid, subc_opts_edit[s_sel], sec_opts[sec_sel] if sec_opts else None, op, of, pn, hr, ce, obs_av)
                            st.success("Avance actualizado")
                            del st.session_state.edit_av
                            st.rerun()

                if st.session_state.get("del_av") == aid:
                    confirm_dialog("del_av", aid, delete_avance)
    else:
        st.info("No hay avances registrados para este día")

    st.markdown("<br>", unsafe_allow_html=True)
    back_button()


# ═══════════════════════════════════════════════
# PAGE: AVANCE GENERAL
# ═══════════════════════════════════════════════
def page_avance_gral():
    show_header()
    st.markdown("## 📈 Avance General de Partidas")

    partidas = get_partidas()
    if not partidas:
        st.info("No hay partidas registradas")
        if st.button("⬅️ Volver"):
            st.session_state.page = "inicio"
            st.rerun()
        return

    fechas = get_registros_fechas()
    if not fechas:
        st.info("Aún no hay registros diarios")
        if st.button("⬅️ Volver"):
            st.session_state.page = "inicio"
            st.rerun()
        return

    for p in partidas:
        total_ejecutado = 0.0
        total_horas = 0.0
        total_operarios = 0
        total_oficiales = 0
        total_peones = 0

        for f in fechas:
            reg = get_registro_diario(f)
            if reg:
                avances = get_avances(reg["id"])
                for a in avances:
                    if a["partida_id"] == p["id"]:
                        total_ejecutado += a["cantidad_ejecutada"]
                        total_horas += a["horas_trabajadas"]
                        total_operarios += a["num_operarios"]
                        total_oficiales += a["num_oficiales"]
                        total_peones += a["num_peones"]

        pct = (total_ejecutado / p["metrado_total"] * 100) if p["metrado_total"] > 0 else 0
        card(
            f"<strong>{p['codigo'] + ' - ' if p['codigo'] else ''}{p['nombre']}</strong>"
            f"<br>📊 Avance: {total_ejecutado:.2f} / {p['metrado_total']:.2f} {p['unidad'] or ''} "
            f"({pct:.1f}%)"
            f"<br>👥 Op:{total_operarios} Of:{total_oficiales} Pn:{total_peones} | ⏱ {total_horas:.1f}h"
            f"<br><progress value='{pct/100}' max='1' style='width:100%;height:12px;border-radius:6px;'></progress>"
        )

    back_button()


# ═══════════════════════════════════════════════
# PAGE: MATERIALES
# ═══════════════════════════════════════════════
def page_materiales():
    show_header()
    st.markdown("## 📦 Control de Materiales")

    catalogo = get_materiales_catalogo()
    if not catalogo:
        st.warning("No hay materiales en el catálogo. Ve a ⚙️ Administrar para agregarlos.")
    else:
        with st.form("form_material"):
            fecha = st.date_input("Fecha", value=date.today())
            mat_opts = {m["nombre"]: m["id"] for m in catalogo}
            mat_sel = st.selectbox("Material", options=list(mat_opts.keys()))
            cantidad = st.number_input("Cantidad", min_value=0.0, value=0.0, step=0.5)
            if st.form_submit_button("💾 Guardar Material"):
                cat_id = mat_opts[mat_sel]
                unidad = next(m["unidad"] for m in catalogo if m["id"] == cat_id)
                save_material(fecha.isoformat(), mat_sel, cantidad, unidad)
                st.success(f"{mat_sel}: {cantidad} registrado")
                st.rerun()

    st.markdown("#### Materiales registrados")
    mats = get_materials(st.session_state.current_fecha)
    if mats:
        for m in mats:
            mid = m["id"]
            card(f"<strong>{m['material']}</strong>: {m['cantidad']} {m['unidad']}")
            cols = st.columns([1, 1, 8])
            with cols[0]:
                if st.button("✏️", key=f"edit_mat_{mid}"):
                    st.session_state.edit_mat = mid
                    st.rerun()
            with cols[1]:
                if st.button("🗑️", key=f"del_mat_{mid}"):
                    st.session_state.del_mat = mid
                    st.rerun()
            if st.session_state.get("edit_mat") == mid:
                with st.form(key=f"form_edit_mat_{mid}"):
                    cat_opts = {m2["nombre"]: m2["id"] for m2 in catalogo}
                    nm = st.selectbox("Material", options=list(cat_opts.keys()), index=list(cat_opts.keys()).index(m["material"]) if m["material"] in cat_opts else 0)
                    ct = st.number_input("Cantidad", min_value=0.0, value=float(m["cantidad"]), step=0.5)
                    if st.form_submit_button("💾 Guardar"):
                        update_material(mid, nm, ct, next(m2["unidad"] for m2 in catalogo if m2["nombre"] == nm))
                        st.success("Actualizado")
                        del st.session_state.edit_mat
                        st.rerun()
            if st.session_state.get("del_mat") == mid:
                confirm_dialog("del_mat", mid, delete_material)
    else:
        st.info("No hay materiales registrados para esta fecha")

    back_button()


# ═══════════════════════════════════════════════
# PAGE: NOTAS DE CAMPO
# ═══════════════════════════════════════════════
def page_notas():
    show_header()
    st.markdown("## 📝 Notas de Campo")

    with st.form("form_nota"):
        fecha = st.date_input("Fecha", value=date.today())
        nota = st.text_area(
            "Describe el incidente o detalle del día:",
            placeholder="Ej: Se detectó fisuras en la viga del eje B-2...",
            height=120,
        )
        if st.form_submit_button("💾 Guardar Nota"):
            if nota.strip():
                save_field_note(fecha.isoformat(), nota.strip())
                st.success("Nota guardada")
                st.rerun()
            else:
                st.warning("Escribe algo antes de guardar")

    st.markdown("#### Notas registradas")
    notas = get_field_notes(st.session_state.current_fecha)
    if notas:
        for n in notas:
            nid = n["id"]
            card(f"📌 {n['nota']}")
            cols = st.columns([1, 1, 8])
            with cols[0]:
                if st.button("✏️", key=f"edit_nota_{nid}"):
                    st.session_state.edit_nota = nid
                    st.rerun()
            with cols[1]:
                if st.button("🗑️", key=f"del_nota_{nid}"):
                    st.session_state.del_nota = nid
                    st.rerun()
            if st.session_state.get("edit_nota") == nid:
                with st.form(key=f"form_edit_nota_{nid}"):
                    txt = st.text_area("Nota", value=n["nota"], height=100)
                    if st.form_submit_button("💾 Guardar"):
                        update_field_note(nid, txt.strip())
                        st.success("Actualizada")
                        del st.session_state.edit_nota
                        st.rerun()
            if st.session_state.get("del_nota") == nid:
                confirm_dialog("del_nota", nid, delete_field_note)
    else:
        st.info("No hay notas para esta fecha")

    back_button()


# ═══════════════════════════════════════════════
# PAGE: FOTOS
# ═══════════════════════════════════════════════
def page_fotos():
    show_header()
    st.markdown("## 📸 Registro Fotográfico")

    with st.form("form_foto"):
        col1, col2 = st.columns(2)
        with col1:
            fecha = st.date_input("Fecha", value=date.today())
        with col2:
            sectores = get_sectores()
            sec_opts = [s["nombre"] for s in sectores]
            sector = st.selectbox("Sector", sec_opts if sec_opts else ["General"])

        descripcion = st.text_input("Descripción breve", placeholder="Ej: Vaciado de techo Bloque A")
        foto = st.file_uploader("Seleccionar foto", type=["jpg", "jpeg", "png", "webp"])

        if st.form_submit_button("💾 Subir Foto"):
            if foto is not None:
                ext = os.path.splitext(foto.name)[1]
                fecha_str = fecha.isoformat()
                sector_short = sector.split(" - ")[0].replace(" ", "_")
                filename = f"{fecha_str}_{sector_short}_{datetime.now().strftime('%H%M%S')}{ext}"
                ruta = os.path.join(PHOTOS_DIR, filename)
                with open(ruta, "wb") as f:
                    f.write(foto.getbuffer())
                save_photo(filename, fecha_str, sector, descripcion, ruta)
                st.success(f"Foto guardada: {filename}")
                st.rerun()
            else:
                st.warning("Selecciona una foto primero")

    st.markdown("#### Fotos registradas")
    fotos = get_photos(st.session_state.current_fecha)
    if fotos:
        for f in fotos:
            pid = f["id"]
            card(
                f"<strong>{f['filename']}</strong>"
                f"{'<br><em>' + f['descripcion'] + '</em>' if f['descripcion'] else ''}"
                f"{'<br>Sector: ' + f['sector'] if f['sector'] else ''}"
            )
            cols = st.columns([1, 1, 8])
            with cols[0]:
                if st.button("✏️", key=f"edit_foto_{pid}"):
                    st.session_state.edit_foto = pid
                    st.rerun()
            with cols[1]:
                if st.button("🗑️", key=f"del_foto_{pid}"):
                    st.session_state.del_foto = pid
                    st.rerun()
            if st.session_state.get("edit_foto") == pid:
                with st.form(key=f"form_edit_foto_{pid}"):
                    sec_opts = [s["nombre"] for s in get_sectores()]
                    sec = st.selectbox("Sector", sec_opts, index=sec_opts.index(f["sector"]) if f["sector"] in sec_opts else 0)
                    desc = st.text_input("Descripción", value=f["descripcion"] or "")
                    if st.form_submit_button("💾 Guardar"):
                        update_photo(pid, sec, desc)
                        st.success("Actualizada")
                        del st.session_state.edit_foto
                        st.rerun()
            if st.session_state.get("del_foto") == pid:
                confirm_dialog("del_foto", pid, delete_photo)
            if os.path.exists(f["ruta"]):
                st.image(f["ruta"], use_container_width=True)
    else:
        st.info("No hay fotos para esta fecha")

    back_button()


# ═══════════════════════════════════════════════
# PAGE: INFORME
# ═══════════════════════════════════════════════
def page_informe():
    show_header()
    st.markdown("## 📄 Generar Reporte Diario de Obra")

    fecha = st.date_input("Selecciona la fecha del reporte", value=date.today())
    fecha_str = fecha.isoformat()

    registro = get_registro_diario(fecha_str)
    if not registro:
        st.warning("No hay registro diario para esta fecha. Agrega uno primero en Registro Diario.")
    else:
        avances = get_avances(registro["id"])
        mats = get_materials(fecha_str)
        notas = get_field_notes(fecha_str)
        fotos = get_photos(fecha_str)

        asunto = st.text_input("ASUNTO", placeholder="Ej: FALTA DE ABASTECIMIENTO DE AGREGADOS")

        st.markdown(
            f"**Resumen:** {len(avances)} partida(s), "
            f"{len(mats)} material(es), "
            f"{len(notas)} nota(s), "
            f"{len(fotos)} foto(s)"
        )

        if st.button("📄 Generar Reporte Word", use_container_width=True):
            resp_nombre = get_config("responsable_nombre", "Ing. MANUEL MEJIA DIAZ")
            resp_cargo = get_config("responsable_cargo", "INGENIERO DE PRODUCCION")
            res_nombre = get_config("residente_nombre", "Ing. WILFREDO PORTAL IDRUGO")
            res_cargo = get_config("residente_cargo", "RESIDENTE DE OBRA")
            proy_nombre = get_config("proyecto_nombre", "")
            cui = get_config("cui", "2336443")
            doc = generar_informe(fecha_str, avances, mats, notas, fotos, registro,
                                  resp_nombre, resp_cargo,
                                  res_nombre, res_cargo,
                                  asunto, proy_nombre, cui)
            import io
            buf = io.BytesIO()
            doc.save(buf)
            buf.seek(0)
            output_name = f"REPORTE_DIARIO_{fecha_str}.docx"
            st.download_button(
                label="⬇️ Descargar Reporte",
                data=buf,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                use_container_width=True,
            )
            try:
                output_path = os.path.join(os.path.dirname(__file__), output_name)
                with open(output_path, "wb") as f:
                    f.write(buf.getvalue())
                st.success(f"Reporte guardado: {output_name}")
            except PermissionError:
                st.info(f"Reporte listo para descargar. Cierra el archivo si está abierto en Word.")

    back_button()


# ═══════════════════════════════════════════════
# PAGE: INFORME SEMANAL
# ═══════════════════════════════════════════════
def page_informe_semanal():
    show_header()
    st.markdown("## 📅 Informe Semanal Consolidado")

    col1, col2 = st.columns(2)
    with col1:
        fecha_ini = st.date_input("Fecha inicio", value=date.today(), key="sem_ini")
    with col2:
        fecha_fin = st.date_input("Fecha fin", value=date.today(), key="sem_fin")

    if fecha_ini > fecha_fin:
        st.error("La fecha inicio debe ser anterior a la fecha fin")
        back_button()
        return

    fecha_ini_str = fecha_ini.isoformat()
    fecha_fin_str = fecha_fin.isoformat()

    registros = get_registros_por_rango(fecha_ini_str, fecha_fin_str)
    avances = get_avances_por_rango(fecha_ini_str, fecha_fin_str)
    materiales = get_materials_por_rango(fecha_ini_str, fecha_fin_str)
    notas = get_field_notes_por_rango(fecha_ini_str, fecha_fin_str)

    st.markdown(
        f"**Resumen:** {len(registros)} día(s), "
        f"{len(avances)} avance(s), "
        f"{len(materiales)} material(es), "
        f"{len(notas)} nota(s)"
    )

    if st.button("📄 Generar Informe Semanal Word", use_container_width=True):
        resp_nombre = get_config("responsable_nombre", "Ing. MANUEL MEJIA DIAZ")
        resp_cargo = get_config("responsable_cargo", "INGENIERO DE PRODUCCION")
        res_nombre = get_config("residente_nombre", "Ing. WILFREDO PORTAL IDRUGO")
        res_cargo = get_config("residente_cargo", "RESIDENTE DE OBRA")
        proy_nombre = get_config("proyecto_nombre", "")
        cui = get_config("cui", "2336443")
        doc = generar_informe_semanal(fecha_ini_str, fecha_fin_str, registros, avances, materiales, notas,
                                      resp_nombre, resp_cargo,
                                      res_nombre, res_cargo,
                                      "", proy_nombre, cui)
        import io
        buf = io.BytesIO()
        doc.save(buf)
        buf.seek(0)
        output_name = f"INFORME_SEMANAL_{fecha_ini_str}_al_{fecha_fin_str}.docx"
        st.download_button(
            label="⬇️ Descargar Informe Semanal",
            data=buf,
            file_name=output_name,
            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            use_container_width=True,
        )
        try:
            output_path = os.path.join(os.path.dirname(__file__), output_name)
            with open(output_path, "wb") as f:
                f.write(buf.getvalue())
            st.success(f"Informe semanal guardado: {output_name}")
        except PermissionError:
            st.info(f"Informe listo para descargar. Cierra el archivo si está abierto en Word.")

    if registros:
        st.divider()
        st.markdown("### 📋 Vista previa por día")
        for r in registros:
            with st.expander(f"📅 {r['fecha']}"):
                st.markdown(f"**Clima:** {r['clima']}")
                avs = [a for a in avances if a["fecha"] == r["fecha"]]
                for a in avs:
                    total_p = a["num_operarios"] + a["num_oficiales"] + a["num_peones"]
                    st.markdown(
                        f"- **{a['partida_codigo'] + ' - ' if a['partida_codigo'] else ''}{a['partida_nombre']}**"
                        f" | 🏗️ {a['sector_nombre'] or 'N/A'}"
                        f" | 🏢 {a['subcontrata_nombre'] or 'Empresa'}"
                        f" | 👥 Op:{a['num_operarios']} Of:{a['num_oficiales']} Pn:{a['num_peones']}"
                        f" | ⏱ {a['horas_trabajadas']}h"
                        f" | 📊 {a['cantidad_ejecutada']} {a['partida_unidad'] or ''}"
                    )

    back_button()


# ═══════════════════════════════════════════════
# PAGE: HISTORIAL
# ═══════════════════════════════════════════════
def page_historial():
    show_header()
    st.markdown("## 📊 Historial de Obra")

    fechas = get_registros_fechas()
    if not fechas:
        fechas = get_all_fechas()

    if not fechas:
        st.info("Aún no hay registros")
        back_button()
        return

    fecha_sel = st.selectbox("Seleccionar fecha", fechas)
    registro = get_registro_diario(fecha_sel)

    if registro:
        rid = registro["id"]
        col1, col2 = st.columns([6, 1])
        with col1:
            clima_edit = st.select_slider("Clima", CLIMA_OPTS, value=registro["clima"], key=f"hist_clima_{rid}")
            obs_edit = st.text_input("Observaciones", value=registro.get("observaciones", ""), key=f"hist_obs_{rid}")
        with col2:
            st.write("")
            st.write("")
            if st.button("💾", key=f"save_hist_reg_{rid}", use_container_width=True):
                conn = get_connection()
                conn.execute("UPDATE registro_diario SET clima=?, observaciones=? WHERE id=?", (clima_edit, obs_edit, rid))
                conn.commit()
                conn.close()
                st.success("Guardado")
                st.rerun()
            if st.button("📋", key=f"goto_reg_{rid}", use_container_width=True):
                st.session_state.current_fecha = fecha_sel
                st.session_state.page = "registro"
                st.rerun()

        avances = get_avances(rid)
        if avances:
            st.markdown("#### 📋 Avances")
            sectores = get_sectores()
            sec_opts = {s["nombre"]: s["id"] for s in sectores}
            subcontratas = get_subcontratas()
            subc_opts = {s["nombre"]: s["id"] for s in subcontratas}
            subc_opts["(De la empresa)"] = None

            for a in avances:
                aid = a["id"]
                total_p = a["num_operarios"] + a["num_oficiales"] + a["num_peones"]
                card(
                    f"<strong>{a['partida_codigo'] + ' - ' if a['partida_codigo'] else ''}{a['partida_nombre']}</strong>"
                    f"<br>🏗️ {a['sector_nombre'] or 'Sin sector'} | 🏢 {a['subcontrata_nombre'] or 'De la empresa'}"
                    f"<br>👥 {total_p} pers | ⏱ {a['horas_trabajadas']}h "
                    f"| 📊 {a['cantidad_ejecutada']} {a['partida_unidad'] or ''}"
                )
                cols = st.columns([1, 1, 8])
                with cols[0]:
                    if st.button("✏️", key=f"hist_edit_av_{aid}", use_container_width=True):
                        st.session_state.hist_edit_av = aid
                        st.rerun()
                with cols[1]:
                    if st.button("🗑️", key=f"hist_del_av_{aid}", use_container_width=True):
                        st.session_state.hist_del_av = aid
                        st.rerun()

                if st.session_state.get("hist_edit_av") == aid:
                    with st.form(key=f"hist_form_edit_av_{aid}"):
                        col1, col2 = st.columns(2)
                        with col1:
                            subc_opts_e = {s["nombre"]: s["id"] for s in get_subcontratas()}
                            subc_opts_e["(De la empresa)"] = None
                            cur_subc = a["subcontrata_nombre"] or "(De la empresa)"
                            s_sel = st.selectbox("Ejecutado por", options=list(subc_opts_e.keys()), index=list(subc_opts_e.keys()).index(cur_subc) if cur_subc in subc_opts_e else 0, key=f"hist_e_subc_{aid}")
                        with col2:
                            cur_sec = a["sector_nombre"] or list(sec_opts.keys())[0] if sec_opts else "(Sin sector)"
                            sec_sel = st.selectbox("Sector", options=list(sec_opts.keys()) if sec_opts else ["(Sin sector)"], index=list(sec_opts.keys()).index(cur_sec) if cur_sec in sec_opts else 0, key=f"hist_e_sec_{aid}")
                        c1, c2, c3 = st.columns(3)
                        with c1:
                            op = st.number_input("Operarios", min_value=0, value=a["num_operarios"], step=1, key=f"hist_e_op_{aid}")
                        with c2:
                            of = st.number_input("Oficiales", min_value=0, value=a["num_oficiales"], step=1, key=f"hist_e_of_{aid}")
                        with c3:
                            pn = st.number_input("Peones", min_value=0, value=a["num_peones"], step=1, key=f"hist_e_pn_{aid}")
                        c1, c2 = st.columns(2)
                        with c1:
                            hr = st.number_input("Horas", min_value=0.0, value=float(a["horas_trabajadas"]), step=0.5, key=f"hist_e_hr_{aid}")
                        with c2:
                            ce = st.number_input("Cant. ejecutada", min_value=0.0, value=float(a["cantidad_ejecutada"]), step=0.1, key=f"hist_e_ce_{aid}")
                        obs_av = st.text_input("Observaciones", value=a["observaciones"] or "", key=f"hist_e_obs_{aid}")
                        if st.form_submit_button("💾 Guardar"):
                            update_avance(aid, subc_opts_e[s_sel], sec_opts[sec_sel] if sec_opts else None, op, of, pn, hr, ce, obs_av)
                            st.success("Avance actualizado")
                            del st.session_state.hist_edit_av
                            st.rerun()

                if st.session_state.get("hist_del_av") == aid:
                    confirm_dialog("hist_del_av", aid, delete_avance)

    notas = get_field_notes(fecha_sel)
    if notas:
        st.markdown("#### 📝 Notas")
        for n in notas:
            nid = n["id"]
            card(f"📌 {n['nota']}")
            cols = st.columns([1, 1, 8])
            with cols[0]:
                if st.button("✏️", key=f"hist_edit_nota_{nid}", use_container_width=True):
                    st.session_state.hist_edit_nota = nid
                    st.rerun()
            with cols[1]:
                if st.button("🗑️", key=f"hist_del_nota_{nid}", use_container_width=True):
                    st.session_state.hist_del_nota = nid
                    st.rerun()
            if st.session_state.get("hist_edit_nota") == nid:
                with st.form(key=f"hist_form_edit_nota_{nid}"):
                    txt = st.text_area("Nota", value=n["nota"], height=100)
                    if st.form_submit_button("💾 Guardar"):
                        update_field_note(nid, txt.strip())
                        st.success("Actualizada")
                        del st.session_state.hist_edit_nota
                        st.rerun()
            if st.session_state.get("hist_del_nota") == nid:
                confirm_dialog("hist_del_nota", nid, delete_field_note)

    fotos = get_photos(fecha_sel)
    if fotos:
        st.markdown("#### 📸 Fotos")
        for f in fotos:
            pid = f["id"]
            if os.path.exists(f["ruta"]):
                st.image(f["ruta"], use_container_width=True)
                st.caption(f"{f['filename']} - {f['descripcion'] or ''}")
            cols = st.columns([1, 1, 8])
            with cols[0]:
                if st.button("✏️", key=f"hist_edit_foto_{pid}", use_container_width=True):
                    st.session_state.hist_edit_foto = pid
                    st.rerun()
            with cols[1]:
                if st.button("🗑️", key=f"hist_del_foto_{pid}", use_container_width=True):
                    st.session_state.hist_del_foto = pid
                    st.rerun()
            if st.session_state.get("hist_edit_foto") == pid:
                with st.form(key=f"hist_form_edit_foto_{pid}"):
                    sec_opts = [s["nombre"] for s in get_sectores()]
                    sec = st.selectbox("Sector", sec_opts, index=sec_opts.index(f["sector"]) if f["sector"] in sec_opts else 0)
                    desc = st.text_input("Descripción", value=f["descripcion"] or "")
                    if st.form_submit_button("💾 Guardar"):
                        update_photo(pid, sec, desc)
                        st.success("Actualizada")
                        del st.session_state.hist_edit_foto
                        st.rerun()
            if st.session_state.get("hist_del_foto") == pid:
                confirm_dialog("hist_del_foto", pid, delete_photo)

    back_button()


# ═══════════════════════════════════════════════
# PAGE: ADMINISTRAR
# ═══════════════════════════════════════════════
def page_admin():
    show_header()
    st.markdown("## ⚙️ Administrar Catálogos")

    tab1, tab2, tab3, tab4, tab5 = st.tabs(["🏗️ Sectores", "📦 Materiales", "📋 Partidas", "👥 Subcontratas", "⚙️ Configuración"])

    # ── TAB 1: SECTORES ──
    with tab1:
        st.markdown("### Sectores de obra")
        with st.form("form_sector"):
            nombre = st.text_input("Nuevo sector", placeholder="Ej: Bloque C - Cimentación")
            if st.form_submit_button("➕ Agregar"):
                if nombre.strip():
                    try:
                        save_sector(nombre.strip())
                        st.success(f"Sector '{nombre}' agregado")
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))

        sectores = get_sectores()
        for s in sectores:
            sid = s["id"]
            card(f"<strong>{s['nombre']}</strong>")
            cols = st.columns([1, 1, 8])
            with cols[0]:
                if st.button("✏️", key=f"edit_sec_{sid}"):
                    st.session_state.edit_sec = sid
                    st.rerun()
            with cols[1]:
                if st.button("🗑️", key=f"del_sec_{sid}", use_container_width=True):
                    st.session_state.del_sec = sid
                    st.rerun()
            if st.session_state.get("edit_sec") == sid:
                with st.form(key=f"form_edit_sec_{sid}"):
                    nm = st.text_input("Nombre", value=s["nombre"])
                    if st.form_submit_button("💾 Guardar"):
                        update_sector(sid, nm.strip())
                        st.success("Actualizado")
                        del st.session_state.edit_sec
                        st.rerun()
            if st.session_state.get("del_sec") == sid:
                confirm_dialog("del_sec", sid, delete_sector)

    # ── TAB 2: MATERIALES ──
    with tab2:
        st.markdown("### Catálogo de Materiales")
        with st.form("form_mat_cat"):
            col1, col2 = st.columns(2)
            with col1:
                nombre = st.text_input("Nombre del material", placeholder="Ej: Cemento Portland")
            with col2:
                unidad = st.text_input("Unidad", placeholder="Ej: bolsas, m3, kg")
            if st.form_submit_button("➕ Agregar"):
                if nombre.strip() and unidad.strip():
                    try:
                        save_material_catalogo(nombre.strip(), unidad.strip())
                        st.success(f"Material '{nombre}' agregado")
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))

        st.markdown("#### Importar desde Excel")
        with st.container():
            archivo_mats = st.file_uploader("Subir archivo Excel (.xlsx)", type=["xlsx"], key="upload_materiales")
            reemplazar_mats = st.checkbox("Reemplazar todos los materiales existentes", key="reemplazar_mats")
            if archivo_mats and st.button("📥 Importar Materiales", key="btn_import_mats"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(archivo_mats)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    col_idx = {h.strip().upper() if h else "": i for i, h in enumerate(headers)}
                    required = ["NOMBRE", "UNIDAD"]
                    missing = [r for r in required if r not in col_idx]
                    if missing:
                        st.error(f"Columnas requeridas faltantes: {', '.join(missing)}. Deben ser: Nombre, Unidad")
                    else:
                        if reemplazar_mats:
                            conn = get_connection()
                            conn.execute("DELETE FROM materiales_catalogo")
                            conn.commit()
                            conn.close()
                        count = 0
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            nombre_val = row[col_idx["NOMBRE"]] if col_idx["NOMBRE"] < len(row) else None
                            unidad_val = row[col_idx["UNIDAD"]] if col_idx["UNIDAD"] < len(row) else None
                            if not nombre_val or not unidad_val:
                                continue
                            save_material_catalogo(str(nombre_val).strip(), str(unidad_val).strip())
                            count += 1
                        st.success(f"{count} materiales importados correctamente")
                        st.rerun()
                except Exception as e:
                    st.error(f"Error al importar: {e}")

        mats = get_materiales_catalogo()
        for m in mats:
            mid = m["id"]
            card(f"<strong>{m['nombre']}</strong> — {m['unidad']}")
            cols = st.columns([1, 1, 8])
            with cols[0]:
                if st.button("✏️", key=f"edit_mat_cat_{mid}"):
                    st.session_state.edit_mat_cat = mid
                    st.rerun()
            with cols[1]:
                if st.button("🗑️", key=f"del_mat_cat_{mid}"):
                    st.session_state.del_mat_cat = mid
                    st.rerun()
            if st.session_state.get("edit_mat_cat") == mid:
                with st.form(key=f"form_edit_mat_cat_{mid}"):
                    c1, c2 = st.columns(2)
                    with c1:
                        nm = st.text_input("Nombre", value=m["nombre"])
                    with c2:
                        un = st.text_input("Unidad", value=m["unidad"])
                    if st.form_submit_button("💾 Guardar"):
                        update_material_catalogo(mid, nm.strip(), un.strip())
                        st.success("Actualizado")
                        del st.session_state.edit_mat_cat
                        st.rerun()
            if st.session_state.get("del_mat_cat") == mid:
                confirm_dialog("del_mat_cat", mid, delete_material_catalogo)

    # ── TAB 3: PARTIDAS ──
    with tab3:
        st.markdown("### Partidas")
        with st.form("form_partida"):
            col1, col2 = st.columns([1, 3])
            with col1:
                codigo = st.text_input("Código", placeholder="Ej: 01.01.01")
            with col2:
                nombre = st.text_input("Nombre de la partida", placeholder="Ej: Muros de ladrillo")
            descripcion = st.text_area("Descripción", placeholder="Descripción detallada de la partida...", height=80)
            col1, col2 = st.columns(2)
            with col1:
                unidad = st.text_input("Unidad", placeholder="Ej: m2, m3, kg")
            with col2:
                metrado = st.number_input("Metrado total", min_value=0.0, value=0.0, step=0.1)
            if st.form_submit_button("➕ Agregar Partida"):
                save_partida(codigo, nombre, descripcion, None, unidad, metrado)
                st.success(f"Partida '{nombre}' agregada")
                st.rerun()

        st.markdown("#### Importar desde Excel")
        with st.container():
            archivo_partidas = st.file_uploader("Subir archivo Excel (.xlsx)", type=["xlsx"], key="upload_partidas")
            reemplazar_partidas = st.checkbox("Reemplazar todas las partidas existentes", key="reemplazar_partidas")
            if archivo_partidas and st.button("📥 Importar Partidas", key="btn_import_partidas"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(archivo_partidas)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    col_idx = {h.strip().upper() if h else "": i for i, h in enumerate(headers)}
                    required = ["NOMBRE"]
                    missing = [r for r in required if r not in col_idx]
                    if missing:
                        st.error(f"Columnas requeridas faltantes: {', '.join(missing)}. Deben ser: Código, Nombre, Descripción, Unidad, Metrado")
                    else:
                        if reemplazar_partidas:
                            conn = get_connection()
                            conn.execute("DELETE FROM partidas")
                            conn.commit()
                            conn.close()
                        count = 0
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            nombre_val = row[col_idx.get("NOMBRE")] if col_idx.get("NOMBRE") is not None and col_idx["NOMBRE"] < len(row) else None
                            if not nombre_val:
                                continue
                            cod = str(row[col_idx["CODIGO"]]) if "CODIGO" in col_idx and col_idx["CODIGO"] < len(row) and row[col_idx["CODIGO"]] is not None else ""
                            desc = str(row[col_idx["DESCRIPCION"]]) if "DESCRIPCION" in col_idx and col_idx["DESCRIPCION"] < len(row) and row[col_idx["DESCRIPCION"]] is not None else ""
                            uni = str(row[col_idx["UNIDAD"]]) if "UNIDAD" in col_idx and col_idx["UNIDAD"] < len(row) and row[col_idx["UNIDAD"]] is not None else ""
                            met = float(row[col_idx["METRADO"]]) if "METRADO" in col_idx and col_idx["METRADO"] < len(row) and row[col_idx["METRADO"]] is not None else 0.0
                            save_partida(cod, str(nombre_val).strip(), desc.strip(), None, uni.strip(), met)
                            count += 1
                        st.success(f"{count} partidas importadas correctamente")
                        st.rerun()
                except Exception as e:
                    st.error(f"Error al importar: {e}")

        partidas = get_partidas()
        for p in partidas:
            pid = p["id"]
            card(
                f"<strong>{p['codigo'] + ' - ' if p['codigo'] else ''}{p['nombre']}</strong>"
                f"<br>{p['descripcion'] or ''}"
                f"<br>Unidad: {p['unidad'] or 'N/A'} | Metrado: {p['metrado_total']}"
            )
            cols = st.columns([1, 1, 8])
            with cols[0]:
                if st.button("✏️", key=f"edit_par_{pid}"):
                    st.session_state.edit_par = pid
                    st.rerun()
            with cols[1]:
                if st.button("🗑️", key=f"del_par_{pid}"):
                    st.session_state.del_par = pid
                    st.rerun()
            if st.session_state.get("edit_par") == pid:
                with st.form(key=f"form_edit_par_{pid}"):
                    c1, c2 = st.columns([1, 3])
                    with c1:
                        cod = st.text_input("Código", value=p["codigo"])
                    with c2:
                        nm = st.text_input("Nombre", value=p["nombre"])
                    desc = st.text_area("Descripción", value=p["descripcion"] or "", height=80)
                    c1, c2 = st.columns(2)
                    with c1:
                        un = st.text_input("Unidad", value=p["unidad"] or "")
                    with c2:
                        mt = st.number_input("Metrado total", min_value=0.0, value=float(p["metrado_total"]), step=0.1)
                    if st.form_submit_button("💾 Guardar"):
                        update_partida(pid, cod, nm, desc, None, un, mt)
                        st.success("Actualizada")
                        del st.session_state.edit_par
                        st.rerun()
            if st.session_state.get("del_par") == pid:
                confirm_dialog("del_par", pid, delete_partida)

    # ── TAB 4: SUBCONTRATAS ──
    with tab4:
        st.markdown("### Subcontratas")
        with st.form("form_subc"):
            col1, col2 = st.columns(2)
            with col1:
                nombre = st.text_input("Nombre de la subcontrata", placeholder="Ej: Construcciones XYZ")
            with col2:
                responsable = st.text_input("Responsable", placeholder="Ej: Juan Pérez")
            telefono = st.text_input("Teléfono", placeholder="Ej: 999 888 777")
            if st.form_submit_button("➕ Agregar"):
                if nombre.strip():
                    save_subcontrata(nombre.strip(), responsable, telefono)
                    st.success(f"Subcontrata '{nombre}' agregada")
                    st.rerun()

        subcs = get_subcontratas()
        for s in subcs:
            sid = s["id"]
            card(
                f"<strong>{s['nombre']}</strong>"
                f"<br>Responsable: {s['responsable'] or 'N/A'} | Tel: {s['telefono'] or 'N/A'}"
            )
            cols = st.columns([1, 1, 8])
            with cols[0]:
                if st.button("✏️", key=f"edit_subc_{sid}"):
                    st.session_state.edit_subc = sid
                    st.rerun()
            with cols[1]:
                if st.button("🗑️", key=f"del_subc_{sid}"):
                    st.session_state.del_subc = sid
                    st.rerun()
            if st.session_state.get("edit_subc") == sid:
                with st.form(key=f"form_edit_subc_{sid}"):
                    nm = st.text_input("Nombre", value=s["nombre"])
                    resp = st.text_input("Responsable", value=s["responsable"] or "")
                    tel = st.text_input("Teléfono", value=s["telefono"] or "")
                    if st.form_submit_button("💾 Guardar"):
                        update_subcontrata(sid, nm.strip(), resp, tel)
                        st.success("Actualizada")
                        del st.session_state.edit_subc
                        st.rerun()
            if st.session_state.get("del_subc") == sid:
                confirm_dialog("del_subc", sid, delete_subcontrata)

    # ── TAB 5: CONFIGURACION ──
    with tab5:
        st.markdown("### Configuración del Informe")
        config = get_all_config()
        with st.form("form_config"):
            st.markdown("**DATOS DEL RESPONSABLE (DE)**")
            nombre = st.text_input("Nombre del responsable", value=config.get("responsable_nombre", ""))
            cargo = st.text_input("Cargo del responsable", value=config.get("responsable_cargo", ""))
            st.markdown("**DATOS DEL RESIDENTE (A)**")
            res_nombre = st.text_input("Nombre del residente", value=config.get("residente_nombre", ""))
            res_cargo = st.text_input("Cargo del residente", value=config.get("residente_cargo", ""))
            st.markdown("**DATOS DEL PROYECTO**")
            proy_nombre = st.text_area("Nombre del proyecto", value=config.get("proyecto_nombre", ""))
            cui = st.text_input("CUI", value=config.get("cui", ""))
            if st.form_submit_button("💾 Guardar configuración"):
                set_config("responsable_nombre", nombre)
                set_config("responsable_cargo", cargo)
                set_config("residente_nombre", res_nombre)
                set_config("residente_cargo", res_cargo)
                set_config("proyecto_nombre", proy_nombre)
                set_config("cui", cui)
                st.success("Configuración guardada")
                st.rerun()

    back_button()


# ═══════════════════════════════════════════════
# ROUTER
# ═══════════════════════════════════════════════
pages = {
    "inicio": page_inicio,
    "registro": page_registro,
    "materiales": page_materiales,
    "notas": page_notas,
    "fotos": page_fotos,
    "informe": page_informe,
    "informe_semanal": page_informe_semanal,
    "historial": page_historial,
    "admin": page_admin,
    "avance_gral": page_avance_gral,
}

pages[st.session_state.page]()
