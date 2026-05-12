import os, sys, shutil
from datetime import date, datetime
from kivy.app import App
from kivy.core.window import Window
from kivy.uix.screenmanager import ScreenManager, Screen, SlideTransition
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.slider import Slider
from kivy.uix.checkbox import CheckBox
from kivy.uix.image import Image
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.progressbar import ProgressBar
from kivy.clock import Clock
from kivy.graphics import Color, Rectangle
from kivy.metrics import dp, sp
from kivy.utils import platform
from functools import partial

sys.path.insert(0, os.path.dirname(__file__))
from database import (
    set_db_path, init_db, get_connection,
    get_sectores, save_sector, update_sector, delete_sector,
    get_materiales_catalogo, save_material_catalogo, update_material_catalogo, delete_material_catalogo,
    get_subcontratas, save_subcontrata, update_subcontrata, delete_subcontrata,
    get_partidas, save_partida, update_partida, delete_partida,
    get_registro_diario, get_or_create_registro_diario,
    save_avance, get_avances, get_avance, update_avance, delete_avance,
    get_registros_fechas,
    save_field_note, get_field_notes, update_field_note, delete_field_note,
    save_photo, get_photos, update_photo, delete_photo,
    save_material, get_materials, update_material, delete_material,
    get_all_fechas,
    get_config, set_config, get_all_config,
    get_avances_por_rango, get_registros_por_rango,
    get_materials_por_rango, get_field_notes_por_rango,
)
from export import generar_informe, generar_informe_semanal

CLIMA_OPTS = ["Soleado", "Nublado", "Lluvia", "Lluvia Intensa"]

YELLOW = "#FFCD00"
BLACK = "#101010"
DARK = "#1a1a1a"
GRAY = "#333333"
WHITE = "#FFFFFF"
RED = "#8B0000"

if platform == "android":
    from android.permissions import request_permissions, Permission
    request_permissions([Permission.WRITE_EXTERNAL_STORAGE, Permission.READ_EXTERNAL_STORAGE, Permission.CAMERA])
    from jnius import autoclass
    Environment = autoclass("android.os.Environment")
    EXT_DIR = Environment.getExternalStorageDirectory().getAbsolutePath()
    PHOTOS_BASE = os.path.join(EXT_DIR, "Anotaciones_Obra")
else:
    PHOTOS_BASE = os.path.join(os.path.dirname(__file__), "Anotaciones_Obra")

PHOTOS_DIR = os.path.join(PHOTOS_BASE, "Fotos")
REPORTS_DIR = PHOTOS_BASE
os.makedirs(PHOTOS_DIR, exist_ok=True)
os.makedirs(REPORTS_DIR, exist_ok=True)


def colored_label(text, color=WHITE, bold=False, size=14, halign="left"):
    markup = f"[color={color}]{text}[/color]"
    lbl = Label(text=markup, markup=True, halign=halign, valign="middle", size_hint_y=None)
    lbl.font_size = sp(size)
    lbl.bold = bold
    lbl.bind(texture_size=lbl.setter("size"))
    return lbl


def title_label(text, size=18):
    return colored_label(text, YELLOW, bold=True, size=size, halign="center")


def section_title(text):
    return colored_label(f"[b]{text}[/b]", YELLOW, bold=True, size=15)


def cat_button(text, on_press, height=dp(48), width=None):
    btn = Button(
        text=text,
        size_hint_y=None,
        height=height,
        background_color=[1, 0.804, 0, 1],
        color=[0, 0, 0, 1],
        bold=True,
        font_size=sp(14),
    )
    if width:
        btn.size_hint_x = None
        btn.width = width
    btn.bind(on_press=on_press)
    return btn


def nav_btn(text, screen_name, app):
    return cat_button(text, lambda x: app.switch_screen(screen_name))


def back_btn(app):
    return cat_button("⬅ VOLVER AL INICIO", lambda x: app.switch_screen("menu"))


def small_btn(text, on_press, color=GRAY):
    btn = Button(
        text=text,
        size_hint_y=None,
        height=dp(36),
        size_hint_x=None,
        width=dp(50),
        background_color=[0.267, 0.267, 0.267, 1],
        color=[1, 1, 1, 1],
        font_size=sp(12),
    )
    btn.bind(on_press=on_press)
    return btn


def del_btn(on_press):
    btn = Button(
        text="Eliminar",
        size_hint_y=None,
        height=dp(36),
        background_color=[0.545, 0, 0, 1],
        color=[1, 1, 1, 1],
        font_size=sp(12),
    )
    btn.bind(on_press=on_press)
    return btn


def confirm_popup(title, message, on_confirm):
    content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(15))
    content.add_widget(colored_label(message, WHITE, size=14))
    btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(44))
    btn_si = Button(text="SI", background_color=[1, 0.804, 0, 1], color=[0, 0, 0, 1], bold=True)
    btn_no = Button(text="NO", background_color=[0.3, 0.3, 0.3, 1], color=[1, 1, 1, 1])
    btn_layout.add_widget(btn_si)
    btn_layout.add_widget(btn_no)
    content.add_widget(btn_layout)
    popup = Popup(title=title, content=content, size_hint=[0.8, 0.3])
    btn_si.bind(on_press=lambda x: [on_confirm(), popup.dismiss()])
    btn_no.bind(on_press=lambda x: popup.dismiss())
    popup.open()


def info_popup(title, message):
    content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(15))
    content.add_widget(colored_label(message, WHITE, size=14))
    btn = Button(text="OK", background_color=[1, 0.804, 0, 1], color=[0, 0, 0, 1], bold=True, size_hint_y=None, height=dp(44))
    content.add_widget(btn)
    popup = Popup(title=title, content=content, size_hint=[0.8, 0.3])
    btn.bind(on_press=lambda x: popup.dismiss())
    popup.open()


class BaseScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.app = App.get_running_app()

    def add_header(self, layout):
        layout.add_widget(title_label("CONTROL DE OBRA", size=20))
        with layout.canvas.after:
            Color(1, 0.804, 0, 1)
            Rectangle(pos=[0, layout.height - dp(2)], size=[Window.width, dp(2)])
        layout.add_widget(Label(size_hint_y=None, height=dp(8)))


class MenuScreen(BaseScreen):
    def on_pre_enter(self):
        self.clear_widgets()
        root = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(12), dp(10)])
        root.bind(minimum_height=root.setter("height"))
        self.add_header(root)

        hoy = date.today().isoformat()
        partidas = get_partidas()
        subc = get_subcontratas()
        reg = get_registro_diario(hoy)

        info_grid = GridLayout(cols=4, spacing=dp(6), size_hint_y=None, height=dp(80))
        for txt in [
            f"[color={YELLOW}]HOY[/color]\n{hoy}",
            f"[color={YELLOW}]PARTIDAS[/color]\n{len(partidas)}",
            f"[color={YELLOW}]SUBCONTRATAS[/color]\n{len(subc)}",
            f"[color={YELLOW}]HOY[/color]\n{'Registrado' if reg else 'Pendiente'}",
        ]:
            lbl = Label(text=txt, markup=True, halign="center", valign="middle", size_hint=(1, 1))
            lbl.font_size = sp(11)
            info_grid.add_widget(lbl)
        root.add_widget(info_grid)
        root.add_widget(Label(size_hint_y=None, height=dp(6)))

        btn_data = [
            ("REGISTRO DIARIO", "registro"),
            ("MATERIALES", "materiales"),
            ("NOTAS DE CAMPO", "notas"),
            ("REGISTRO FOTOGRAFICO", "fotos"),
            ("GENERAR INFORME", "informe"),
            ("INFORME SEMANAL", "informe_semanal"),
            ("ADMINISTRAR", "admin"),
            ("AVANCE GENERAL", "avance_gral"),
            ("VER HISTORIAL", "historial"),
        ]
        for text, screen in btn_data:
            btn = cat_button(text, lambda x, s=screen: self.app.switch_screen(s))
            root.add_widget(btn)

        scroll = ScrollView(size_hint=(1, 1))
        scroll.add_widget(root)
        self.add_widget(scroll)


class RegistroScreen(BaseScreen):
    def on_pre_enter(self):
        self.clear_widgets()
        root = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(12), dp(10)])
        root.bind(minimum_height=root.setter("height"))
        self.add_header(root)
        root.add_widget(section_title("REGISTRO DIARIO DE OBRA"))

        self.fecha_input = TextInput(text=date.today().isoformat(), size_hint_y=None, height=dp(40),
                                      foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        root.add_widget(colored_label("Fecha"))
        root.add_widget(self.fecha_input)

        self.clima_spinner = Spinner(text="Soleado", values=CLIMA_OPTS, size_hint_y=None, height=dp(40),
                                      background_color=[0,0,0,1], color=[1,1,1,1])
        root.add_widget(colored_label("Clima"))
        root.add_widget(self.clima_spinner)

        self.obs_input = TextInput(text="", size_hint_y=None, height=dp(40),
                                    foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        root.add_widget(colored_label("Observaciones"))
        root.add_widget(self.obs_input)

        root.add_widget(cat_button("GUARDAR DATOS DEL DIA", self._guardar_dia))
        root.add_widget(Label(size_hint_y=None, height=dp(10)))
        root.add_widget(section_title("AGREGAR AVANCE DE PARTIDA"))

        partidas = get_partidas()
        if not partidas:
            root.add_widget(colored_label("No hay partidas. ADMINISTRAR > Partidas", YELLOW, size=12))
        else:
            self.partida_spinner = Spinner(text="Seleccionar...", values=[f"{p['codigo']+' - ' if p['codigo'] else ''}{p['nombre']}" for p in partidas],
                                            size_hint_y=None, height=dp(40), background_color=[0,0,0,1], color=[1,1,1,1])
            root.add_widget(colored_label("Partida"))
            root.add_widget(self.partida_spinner)

            sectores = get_sectores()
            sec_names = [s["nombre"] for s in sectores] or ["(Sin sector)"]
            self.sector_spinner = Spinner(text=sec_names[0], values=sec_names, size_hint_y=None, height=dp(40),
                                           background_color=[0,0,0,1], color=[1,1,1,1])
            root.add_widget(colored_label("Sector"))
            root.add_widget(self.sector_spinner)

            subcs = get_subcontratas()
            subc_names = [s["nombre"] for s in subcs] + ["(De la empresa)"]
            self.subc_spinner = Spinner(text=subc_names[0], values=subc_names, size_hint_y=None, height=dp(40),
                                         background_color=[0,0,0,1], color=[1,1,1,1])
            root.add_widget(colored_label("Ejecutado por"))
            root.add_widget(self.subc_spinner)

            grid = GridLayout(cols=3, spacing=dp(4), size_hint_y=None, height=dp(80))
            self.op_input = TextInput(text="0", size_hint_y=None, height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
            self.of_input = TextInput(text="0", size_hint_y=None, height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
            self.pn_input = TextInput(text="0", size_hint_y=None, height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
            grid.add_widget(colored_label("Operarios", size=11))
            grid.add_widget(colored_label("Oficiales", size=11))
            grid.add_widget(colored_label("Peones", size=11))
            grid.add_widget(self.op_input)
            grid.add_widget(self.of_input)
            grid.add_widget(self.pn_input)
            root.add_widget(grid)

            grid2 = GridLayout(cols=2, spacing=dp(4), size_hint_y=None, height=dp(80))
            self.hr_input = TextInput(text="0.0", size_hint_y=None, height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
            self.ce_input = TextInput(text="0.0", size_hint_y=None, height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
            grid2.add_widget(colored_label("Horas", size=11))
            grid2.add_widget(colored_label("Cant. Ejec.", size=11))
            grid2.add_widget(self.hr_input)
            grid2.add_widget(self.ce_input)
            root.add_widget(grid2)

            self.av_obs_input = TextInput(text="", size_hint_y=None, height=dp(36),
                                           foreground_color=[1,1,1,1], background_color=[0,0,0,1])
            root.add_widget(colored_label("Observaciones"))
            root.add_widget(self.av_obs_input)
            root.add_widget(cat_button("AGREGAR AVANCE", self._add_avance))

        root.add_widget(Label(size_hint_y=None, height=dp(10)))
        root.add_widget(section_title("AVANCES DEL DIA"))
        self.avances_box = BoxLayout(orientation="vertical", spacing=dp(4), size_hint_y=None)
        self.avances_box.bind(minimum_height=self.avances_box.setter("height"))
        root.add_widget(self.avances_box)

        self._refresh_avances()
        root.add_widget(Label(size_hint_y=None, height=dp(10)))
        root.add_widget(back_btn(self.app))

        scroll = ScrollView(size_hint=(1, 1))
        scroll.add_widget(root)
        self.add_widget(scroll)

    def _get_fecha(self):
        try:
            return self.fecha_input.text.strip()
        except:
            return date.today().isoformat()

    def _guardar_dia(self, *args):
        fecha = self._get_fecha()
        rid = get_or_create_registro_diario(fecha)
        conn = get_connection()
        conn.execute("UPDATE registro_diario SET clima=?, observaciones=? WHERE id=?",
                     (self.clima_spinner.text, self.obs_input.text, rid))
        conn.commit()
        conn.close()
        info_popup("OK", "Datos del dia guardados")

    def _add_avance(self, *args):
        fecha = self._get_fecha()
        rid = get_or_create_registro_diario(fecha)
        partidas = get_partidas()
        sel_part = self.partida_spinner.text
        pid = None
        for p in partidas:
            label = f"{p['codigo']+' - ' if p['codigo'] else ''}{p['nombre']}"
            if label == sel_part:
                pid = p["id"]
                break
        if not pid:
            info_popup("Error", "Selecciona una partida")
            return

        subcs = get_subcontratas()
        subc_id = None
        for s in subcs:
            if s["nombre"] == self.subc_spinner.text:
                subc_id = s["id"]
                break

        sectores = get_sectores()
        sec_id = None
        for s in sectores:
            if s["nombre"] == self.sector_spinner.text:
                sec_id = s["id"]
                break

        try:
            op = int(self.op_input.text)
            of = int(self.of_input.text)
            pn = int(self.pn_input.text)
            hr = float(self.hr_input.text)
            ce = float(self.ce_input.text)
        except:
            info_popup("Error", "Valores numericos invalidos")
            return

        save_avance(rid, pid, subc_id, sec_id, op, of, pn, hr, ce, self.av_obs_input.text)
        info_popup("OK", "Avance agregado")
        self._refresh_avances()

    def _refresh_avances(self):
        self.avances_box.clear_widgets()
        fecha = self._get_fecha()
        reg = get_registro_diario(fecha)
        if not reg:
            self.avances_box.add_widget(colored_label("Sin registro para esta fecha", size=12))
            return
        avances = get_avances(reg["id"])
        if not avances:
            self.avances_box.add_widget(colored_label("No hay avances", size=12))
            return
        for a in avances:
            total = a["num_operarios"] + a["num_oficiales"] + a["num_peones"]
            txt = (f"{a['partida_codigo']+' - ' if a['partida_codigo'] else ''}{a['partida_nombre']}\n"
                   f" Sector: {a['sector_nombre'] or 'N/A'} | {a['subcontrata_nombre'] or 'Empresa'}\n"
                   f" {total} pers | {a['horas_trabajadas']}h | {a['cantidad_ejecutada']} {a['partida_unidad'] or ''}")
            lbl = colored_label(txt, WHITE, size=11)
            lbl.text_size = (Window.width - dp(80), None)
            self.avances_box.add_widget(lbl)
            btn_box = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(8))
            btn_box.add_widget(small_btn("EDIT", lambda x, aid=a["id"]: self._edit_avance(aid)))
            btn_box.add_widget(del_btn(lambda x, aid=a["id"]: self._del_avance(aid)))
            self.avances_box.add_widget(btn_box)

    def _edit_avance(self, aid):
        a = get_avance(aid)
        if not a:
            return
        content = BoxLayout(orientation="vertical", spacing=dp(6), padding=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        op_i = TextInput(text=str(a["num_operarios"]), size_hint_y=None, height=dp(36),
                         foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        of_i = TextInput(text=str(a["num_oficiales"]), size_hint_y=None, height=dp(36),
                         foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        pn_i = TextInput(text=str(a["num_peones"]), size_hint_y=None, height=dp(36),
                         foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        hr_i = TextInput(text=str(a["horas_trabajadas"]), size_hint_y=None, height=dp(36),
                         foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        ce_i = TextInput(text=str(a["cantidad_ejecutada"]), size_hint_y=None, height=dp(36),
                         foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        for lbl, w in [("Operarios", op_i), ("Oficiales", of_i), ("Peones", pn_i),
                        ("Horas", hr_i), ("Cant. Ejec.", ce_i)]:
            content.add_widget(colored_label(lbl, YELLOW, size=12))
            content.add_widget(w)
        btn = cat_button("GUARDAR", lambda x: self._do_edit_avance(aid, op_i, of_i, pn_i, hr_i, ce_i, popup), height=dp(40))
        content.add_widget(btn)
        popup = Popup(title="Editar avance", content=content, size_hint=[0.9, None], height=dp(420))
        popup.open()

    def _do_edit_avance(self, aid, op_i, of_i, pn_i, hr_i, ce_i, popup):
        try:
            op = int(op_i.text)
            of = int(of_i.text)
            pn = int(pn_i.text)
            hr = float(hr_i.text)
            ce = float(ce_i.text)
        except:
            info_popup("Error", "Valores invalidos")
            return
        a = get_avance(aid)
        update_avance(aid, a["subcontrata_id"], a["sector_id"], op, of, pn, hr, ce, a["observaciones"] or "")
        popup.dismiss()
        self._refresh_avances()

    def _del_avance(self, aid):
        confirm_popup("Eliminar", "Eliminar este avance?", lambda: [delete_avance(aid), self._refresh_avances()])


class MaterialesScreen(BaseScreen):
    def on_pre_enter(self):
        self.clear_widgets()
        root = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(12), dp(10)])
        root.bind(minimum_height=root.setter("height"))
        self.add_header(root)
        root.add_widget(section_title("CONTROL DE MATERIALES"))

        catalogo = get_materiales_catalogo()
        if not catalogo:
            root.add_widget(colored_label("Sin catalogo. ADMINISTRAR > Materiales", YELLOW, size=12))
        else:
            self.mat_spinner = Spinner(text=catalogo[0]["nombre"], values=[m["nombre"] for m in catalogo],
                                        size_hint_y=None, height=dp(40), background_color=[0,0,0,1], color=[1,1,1,1])
            root.add_widget(colored_label("Material"))
            root.add_widget(self.mat_spinner)
            self.cant_input = TextInput(text="0.0", size_hint_y=None, height=dp(40),
                                         foreground_color=[1,1,1,1], background_color=[0,0,0,1])
            root.add_widget(colored_label("Cantidad"))
            root.add_widget(self.cant_input)
            root.add_widget(cat_button("GUARDAR MATERIAL", self._guardar_mat))

        root.add_widget(Label(size_hint_y=None, height=dp(8)))
        root.add_widget(section_title("MATERIALES REGISTRADOS"))
        self.mats_box = BoxLayout(orientation="vertical", spacing=dp(4), size_hint_y=None)
        self.mats_box.bind(minimum_height=self.mats_box.setter("height"))
        root.add_widget(self.mats_box)
        self._refresh_mats()
        root.add_widget(Label(size_hint_y=None, height=dp(10)))
        root.add_widget(back_btn(self.app))
        scroll = ScrollView(size_hint=(1, 1))
        scroll.add_widget(root)
        self.add_widget(scroll)

    def _guardar_mat(self, *args):
        catalogo = get_materiales_catalogo()
        cat_id = None
        unidad = ""
        for m in catalogo:
            if m["nombre"] == self.mat_spinner.text:
                cat_id = m["id"]
                unidad = m["unidad"]
                break
        try:
            cant = float(self.cant_input.text)
        except:
            info_popup("Error", "Cantidad invalida")
            return
        save_material(date.today().isoformat(), self.mat_spinner.text, cant, unidad)
        info_popup("OK", f"{self.mat_spinner.text}: {cant} registrado")
        self._refresh_mats()

    def _refresh_mats(self):
        self.mats_box.clear_widgets()
        mats = get_materials(date.today().isoformat())
        if not mats:
            self.mats_box.add_widget(colored_label("No hay materiales hoy", size=12))
            return
        for m in mats:
            lbl = colored_label(f"{m['material']}: {m['cantidad']} {m['unidad']}", WHITE, size=12)
            self.mats_box.add_widget(lbl)
            btn_box = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(8))
            btn_box.add_widget(small_btn("EDIT", lambda x, mid=m["id"]: self._edit_mat(mid)))
            btn_box.add_widget(del_btn(lambda x, mid=m["id"]: self._del_mat(mid)))
            self.mats_box.add_widget(btn_box)

    def _edit_mat(self, mid):
        mats = get_materials(date.today().isoformat())
        m = next((x for x in mats if x["id"] == mid), None)
        if not m:
            return
        content = BoxLayout(orientation="vertical", spacing=dp(6), padding=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        cant_i = TextInput(text=str(m["cantidad"]), size_hint_y=None, height=dp(36),
                            foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        content.add_widget(colored_label("Cantidad", YELLOW, size=12))
        content.add_widget(cant_i)
        btn = cat_button("GUARDAR", lambda x: [update_material(mid, m["material"], float(cant_i.text), m["unidad"]),
                                                popup.dismiss(), self._refresh_mats()], height=dp(40))
        content.add_widget(btn)
        popup = Popup(title="Editar material", content=content, size_hint=[0.9, None], height=dp(200))
        popup.open()

    def _del_mat(self, mid):
        confirm_popup("Eliminar", "Eliminar material?", lambda: [delete_material(mid), self._refresh_mats()])


class NotasScreen(BaseScreen):
    def on_pre_enter(self):
        self.clear_widgets()
        root = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(12), dp(10)])
        root.bind(minimum_height=root.setter("height"))
        self.add_header(root)
        root.add_widget(section_title("NOTAS DE CAMPO"))

        self.nota_input = TextInput(text="", size_hint_y=None, height=dp(100),
                                     foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        root.add_widget(colored_label("Describe el incidente o detalle:"))
        root.add_widget(self.nota_input)
        root.add_widget(cat_button("GUARDAR NOTA", self._guardar_nota))

        root.add_widget(Label(size_hint_y=None, height=dp(8)))
        root.add_widget(section_title("NOTAS REGISTRADAS"))
        self.notas_box = BoxLayout(orientation="vertical", spacing=dp(4), size_hint_y=None)
        self.notas_box.bind(minimum_height=self.notas_box.setter("height"))
        root.add_widget(self.notas_box)
        self._refresh_notas()
        root.add_widget(Label(size_hint_y=None, height=dp(10)))
        root.add_widget(back_btn(self.app))
        scroll = ScrollView(size_hint=(1, 1))
        scroll.add_widget(root)
        self.add_widget(scroll)

    def _guardar_nota(self, *args):
        txt = self.nota_input.text.strip()
        if txt:
            save_field_note(date.today().isoformat(), txt)
            info_popup("OK", "Nota guardada")
            self.nota_input.text = ""
            self._refresh_notas()

    def _refresh_notas(self):
        self.notas_box.clear_widgets()
        notas = get_field_notes(date.today().isoformat())
        if not notas:
            self.notas_box.add_widget(colored_label("No hay notas hoy", size=12))
            return
        for n in notas:
            lbl = colored_label(n["nota"], WHITE, size=12)
            lbl.text_size = (Window.width - dp(80), None)
            self.notas_box.add_widget(lbl)
            btn_box = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(8))
            btn_box.add_widget(small_btn("EDIT", lambda x, nid=n["id"]: self._edit_nota(nid)))
            btn_box.add_widget(del_btn(lambda x, nid=n["id"]: self._del_nota(nid)))
            self.notas_box.add_widget(btn_box)

    def _edit_nota(self, nid):
        notas = get_field_notes(date.today().isoformat())
        n = next((x for x in notas if x["id"] == nid), None)
        if not n:
            return
        content = BoxLayout(orientation="vertical", spacing=dp(6), padding=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        txt_i = TextInput(text=n["nota"], size_hint_y=None, height=dp(100),
                           foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        content.add_widget(colored_label("Nota", YELLOW, size=12))
        content.add_widget(txt_i)
        btn = cat_button("GUARDAR", lambda x: [update_field_note(nid, txt_i.text.strip()),
                                                popup.dismiss(), self._refresh_notas()], height=dp(40))
        content.add_widget(btn)
        popup = Popup(title="Editar nota", content=content, size_hint=[0.9, None], height=dp(280))
        popup.open()

    def _del_nota(self, nid):
        confirm_popup("Eliminar", "Eliminar nota?", lambda: [delete_field_note(nid), self._refresh_notas()])


class FotosScreen(BaseScreen):
    def on_pre_enter(self):
        self.clear_widgets()
        root = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(12), dp(10)])
        root.bind(minimum_height=root.setter("height"))
        self.add_header(root)
        root.add_widget(section_title("REGISTRO FOTOGRAFICO"))

        self.fecha_input = TextInput(text=date.today().isoformat(), size_hint_y=None, height=dp(40),
                                      foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        root.add_widget(colored_label("Fecha"))
        root.add_widget(self.fecha_input)

        sectores = get_sectores()
        sec_names = [s["nombre"] for s in sectores] or ["General"]
        self.sector_spinner = Spinner(text=sec_names[0], values=sec_names, size_hint_y=None, height=dp(40),
                                       background_color=[0,0,0,1], color=[1,1,1,1])
        root.add_widget(colored_label("Sector"))
        root.add_widget(self.sector_spinner)

        self.desc_input = TextInput(text="", size_hint_y=None, height=dp(40),
                                     foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        root.add_widget(colored_label("Descripcion"))
        root.add_widget(self.desc_input)

        root.add_widget(cat_button("TOMAR FOTO", self._tomar_foto))

        root.add_widget(Label(size_hint_y=None, height=dp(8)))
        root.add_widget(section_title("FOTOS REGISTRADAS"))
        self.fotos_box = BoxLayout(orientation="vertical", spacing=dp(4), size_hint_y=None)
        self.fotos_box.bind(minimum_height=self.fotos_box.setter("height"))
        root.add_widget(self.fotos_box)
        self._refresh_fotos()
        root.add_widget(Label(size_hint_y=None, height=dp(10)))
        root.add_widget(back_btn(self.app))
        scroll = ScrollView(size_hint=(1, 1))
        scroll.add_widget(root)
        self.add_widget(scroll)

    def _tomar_foto(self, *args):
        fecha = self.fecha_input.text.strip()
        sector = self.sector_spinner.text
        desc = self.desc_input.text.strip()
        try:
            from plyer import camera
            filename = f"{fecha}_{sector.replace(' ','_')}_{datetime.now().strftime('%H%M%S')}.jpg"
            filepath = os.path.join(PHOTOS_DIR, filename)
            camera.take_picture(filename=filepath, on_complete=lambda p: self._foto_lista(p, fecha, sector, desc, filename))
        except Exception as e:
            info_popup("Error", f"Camara no disponible: {e}")

    def _foto_lista(self, path, fecha, sector, desc, filename):
        if path and os.path.exists(path):
            save_photo(filename, fecha, sector, desc, path)
            info_popup("OK", "Foto guardada")
            self._refresh_fotos()

    def _refresh_fotos(self):
        self.fotos_box.clear_widgets()
        fotos = get_photos(date.today().isoformat())
        if not fotos:
            self.fotos_box.add_widget(colored_label("No hay fotos hoy", size=12))
            return
        for f in fotos:
            lbl = colored_label(f"{f['filename']}\n{f['descripcion'] or ''}", WHITE, size=11)
            lbl.text_size = (Window.width - dp(80), None)
            self.fotos_box.add_widget(lbl)
            if os.path.exists(f["ruta"]):
                img = Image(source=f["ruta"], size_hint_y=None, height=dp(150))
                self.fotos_box.add_widget(img)
            btn_box = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(8))
            btn_box.add_widget(small_btn("EDIT", lambda x, pid=f["id"]: self._edit_foto(pid)))
            btn_box.add_widget(del_btn(lambda x, pid=f["id"]: self._del_foto(pid)))
            self.fotos_box.add_widget(btn_box)

    def _edit_foto(self, pid):
        fotos = get_photos(date.today().isoformat())
        f = next((x for x in fotos if x["id"] == pid), None)
        if not f:
            return
        content = BoxLayout(orientation="vertical", spacing=dp(6), padding=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        desc_i = TextInput(text=f["descripcion"] or "", size_hint_y=None, height=dp(36),
                            foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        content.add_widget(colored_label("Descripcion", YELLOW, size=12))
        content.add_widget(desc_i)
        btn = cat_button("GUARDAR", lambda x: [update_photo(pid, f["sector"], desc_i.text.strip()),
                                                popup.dismiss(), self._refresh_fotos()], height=dp(40))
        content.add_widget(btn)
        popup = Popup(title="Editar foto", content=content, size_hint=[0.9, None], height=dp(200))
        popup.open()

    def _del_foto(self, pid):
        confirm_popup("Eliminar", "Eliminar foto?", lambda: [delete_photo(pid), self._refresh_fotos()])


class InformeScreen(BaseScreen):
    def on_pre_enter(self):
        self.clear_widgets()
        root = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(12), dp(10)])
        root.bind(minimum_height=root.setter("height"))
        self.add_header(root)
        root.add_widget(section_title("GENERAR REPORTE DIARIO"))

        self.fecha_input = TextInput(text=date.today().isoformat(), size_hint_y=None, height=dp(40),
                                      foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        root.add_widget(colored_label("Fecha del reporte"))
        root.add_widget(self.fecha_input)

        self.asunto_input = TextInput(text="", size_hint_y=None, height=dp(40),
                                       foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        root.add_widget(colored_label("ASUNTO (opcional)"))
        root.add_widget(self.asunto_input)

        root.add_widget(cat_button("GENERAR REPORTE WORD", self._generar))
        root.add_widget(Label(size_hint_y=None, height=dp(10)))
        root.add_widget(back_btn(self.app))
        scroll = ScrollView(size_hint=(1, 1))
        scroll.add_widget(root)
        self.add_widget(scroll)

    def _generar(self, *args):
        
        fecha = self.fecha_input.text.strip()
        registro = get_registro_diario(fecha)
        if not registro:
            info_popup("Error", "No hay registro para esta fecha")
            return
        avances = get_avances(registro["id"])
        mats = get_materials(fecha)
        notas = get_field_notes(fecha)
        fotos = get_photos(fecha)
        resp_nombre = get_config("responsable_nombre", "Ing. MANUEL MEJIA DIAZ")
        resp_cargo = get_config("responsable_cargo", "INGENIERO DE PRODUCCION")
        res_nombre = get_config("residente_nombre", "Ing. WILFREDO PORTAL IDRUGO")
        res_cargo = get_config("residente_cargo", "RESIDENTE DE OBRA")
        proy_nombre = get_config("proyecto_nombre", "")
        cui = get_config("cui", "2336443")
        try:
            doc = generar_informe(fecha, avances, mats, notas, fotos, registro,
                                  resp_nombre, resp_cargo, res_nombre, res_cargo,
                                  self.asunto_input.text, proy_nombre, cui)
            out_name = f"REPORTE_DIARIO_{fecha}.docx"
            out_path = os.path.join(REPORTS_DIR, out_name)
            doc.save(out_path)
            info_popup("OK", f"Reporte guardado en:\n{out_path}")
        except Exception as e:
            info_popup("Error", f"Error al generar: {e}")


class InformeSemanalScreen(BaseScreen):
    def on_pre_enter(self):
        self.clear_widgets()
        root = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(12), dp(10)])
        root.bind(minimum_height=root.setter("height"))
        self.add_header(root)
        root.add_widget(section_title("INFORME SEMANAL"))

        self.ini_input = TextInput(text=date.today().isoformat(), size_hint_y=None, height=dp(40),
                                    foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        root.add_widget(colored_label("Fecha inicio"))
        root.add_widget(self.ini_input)
        self.fin_input = TextInput(text=date.today().isoformat(), size_hint_y=None, height=dp(40),
                                    foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        root.add_widget(colored_label("Fecha fin"))
        root.add_widget(self.fin_input)
        root.add_widget(cat_button("GENERAR INFORME SEMANAL", self._generar))
        root.add_widget(Label(size_hint_y=None, height=dp(10)))
        root.add_widget(back_btn(self.app))
        scroll = ScrollView(size_hint=(1, 1))
        scroll.add_widget(root)
        self.add_widget(scroll)

    def _generar(self, *args):
        
        ini = self.ini_input.text.strip()
        fin = self.fin_input.text.strip()
        if ini > fin:
            info_popup("Error", "Fecha inicio debe ser anterior a fin")
            return
        registros = get_registros_por_rango(ini, fin)
        avances = get_avances_por_rango(ini, fin)
        mats = get_materials_por_rango(ini, fin)
        notas = get_field_notes_por_rango(ini, fin)
        resp_nombre = get_config("responsable_nombre", "Ing. MANUEL MEJIA DIAZ")
        resp_cargo = get_config("responsable_cargo", "INGENIERO DE PRODUCCION")
        res_nombre = get_config("residente_nombre", "Ing. WILFREDO PORTAL IDRUGO")
        res_cargo = get_config("residente_cargo", "RESIDENTE DE OBRA")
        proy_nombre = get_config("proyecto_nombre", "")
        cui = get_config("cui", "2336443")
        try:
            doc = generar_informe_semanal(ini, fin, registros, avances, mats, notas,
                                          resp_nombre, resp_cargo, res_nombre, res_cargo,
                                          "", proy_nombre, cui)
            out_name = f"INFORME_SEMANAL_{ini}_al_{fin}.docx"
            out_path = os.path.join(REPORTS_DIR, out_name)
            doc.save(out_path)
            info_popup("OK", f"Informe semanal guardado en:\n{out_path}")
        except Exception as e:
            info_popup("Error", f"Error: {e}")


class HistorialScreen(BaseScreen):
    def on_pre_enter(self):
        self.clear_widgets()
        root = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(12), dp(10)])
        root.bind(minimum_height=root.setter("height"))
        self.add_header(root)
        root.add_widget(section_title("HISTORIAL DE OBRA"))

        fechas = get_registros_fechas() or get_all_fechas()
        if not fechas:
            root.add_widget(colored_label("Sin registros", YELLOW, size=12))
            root.add_widget(back_btn(self.app))
            scroll = ScrollView(size_hint=(1, 1))
            scroll.add_widget(root)
            self.add_widget(scroll)
            return

        self.fecha_spinner = Spinner(text=fechas[0], values=fechas, size_hint_y=None, height=dp(40),
                                      background_color=[0,0,0,1], color=[1,1,1,1])
        self.fecha_spinner.bind(text=self._on_fecha_change)
        root.add_widget(colored_label("Seleccionar fecha"))
        root.add_widget(self.fecha_spinner)

        self.hist_box = BoxLayout(orientation="vertical", spacing=dp(4), size_hint_y=None)
        self.hist_box.bind(minimum_height=self.hist_box.setter("height"))
        root.add_widget(self.hist_box)
        root.add_widget(Label(size_hint_y=None, height=dp(10)))
        root.add_widget(back_btn(self.app))
        scroll = ScrollView(size_hint=(1, 1))
        scroll.add_widget(root)
        self.add_widget(scroll)
        self._load_data()

    def _on_fecha_change(self, spinner, text):
        self._load_data()

    def _load_data(self):
        self.hist_box.clear_widgets()
        fecha = self.fecha_spinner.text
        reg = get_registro_diario(fecha)
        if reg:
            self.hist_box.add_widget(colored_label(f"Clima: {reg['clima']}", YELLOW, size=13))
            if reg.get("observaciones"):
                self.hist_box.add_widget(colored_label(f"Obs: {reg['observaciones']}", WHITE, size=12))
            avances = get_avances(reg["id"])
            if avances:
                self.hist_box.add_widget(colored_label("AVANCES:", YELLOW, bold=True, size=13))
                for a in avances:
                    total = a["num_operarios"] + a["num_oficiales"] + a["num_peones"]
                    txt = (f"{a['partida_codigo']+' - ' if a['partida_codigo'] else ''}{a['partida_nombre']}\n"
                           f" {a['sector_nombre'] or 'N/A'} | Op:{a['num_operarios']} Of:{a['num_oficiales']} Pn:{a['num_peones']}\n"
                           f" {a['horas_trabajadas']}h | {a['cantidad_ejecutada']} {a['partida_unidad'] or ''}")
                    lbl = colored_label(txt, WHITE, size=11)
                    lbl.text_size = (Window.width - dp(80), None)
                    self.hist_box.add_widget(lbl)
                    btn_box = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(8))
                    btn_box.add_widget(small_btn("EDIT", lambda x, aid=a["id"]: self._edit_avance_hist(aid)))
                    btn_box.add_widget(del_btn(lambda x, aid=a["id"]: self._del_avance_hist(aid)))
                    self.hist_box.add_widget(btn_box)

        notas = get_field_notes(fecha)
        if notas:
            self.hist_box.add_widget(colored_label("NOTAS:", YELLOW, bold=True, size=13))
            for n in notas:
                lbl = colored_label(n["nota"], WHITE, size=12)
                lbl.text_size = (Window.width - dp(80), None)
                self.hist_box.add_widget(lbl)
                btn_box = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(8))
                btn_box.add_widget(small_btn("EDIT", lambda x, nid=n["id"]: self._edit_nota_hist(nid)))
                btn_box.add_widget(del_btn(lambda x, nid=n["id"]: self._del_nota_hist(nid)))
                self.hist_box.add_widget(btn_box)

        fotos = get_photos(fecha)
        if fotos:
            self.hist_box.add_widget(colored_label("FOTOS:", YELLOW, bold=True, size=13))
            for f in fotos:
                lbl = colored_label(f"{f['filename']} - {f['descripcion'] or ''}", WHITE, size=11)
                self.hist_box.add_widget(lbl)
                if os.path.exists(f["ruta"]):
                    self.hist_box.add_widget(Image(source=f["ruta"], size_hint_y=None, height=dp(150)))
                btn_box = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(8))
                btn_box.add_widget(small_btn("EDIT", lambda x, pid=f["id"]: self._edit_foto_hist(pid)))
                btn_box.add_widget(del_btn(lambda x, pid=f["id"]: self._del_foto_hist(pid)))
                self.hist_box.add_widget(btn_box)

    def _edit_avance_hist(self, aid):
        a = get_avance(aid)
        if not a:
            return
        content = BoxLayout(orientation="vertical", spacing=dp(6), padding=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        op_i = TextInput(text=str(a["num_operarios"]), height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        of_i = TextInput(text=str(a["num_oficiales"]), height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        pn_i = TextInput(text=str(a["num_peones"]), height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        hr_i = TextInput(text=str(a["horas_trabajadas"]), height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        ce_i = TextInput(text=str(a["cantidad_ejecutada"]), height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        for lbl, w in [("Operarios", op_i), ("Oficiales", of_i), ("Peones", pn_i),
                        ("Horas", hr_i), ("Cant. Ejec.", ce_i)]:
            content.add_widget(colored_label(lbl, YELLOW, size=12))
            content.add_widget(w)
        btn = cat_button("GUARDAR", lambda x: self._do_edit_av_hist(aid, a, op_i, of_i, pn_i, hr_i, ce_i, popup))
        content.add_widget(btn)
        popup = Popup(title="Editar avance", content=content, size_hint=[0.9, None], height=dp(420))
        popup.open()

    def _do_edit_av_hist(self, aid, a, op_i, of_i, pn_i, hr_i, ce_i, popup):
        try:
            update_avance(aid, a["subcontrata_id"], a["sector_id"],
                          int(op_i.text), int(of_i.text), int(pn_i.text),
                          float(hr_i.text), float(ce_i.text), a["observaciones"] or "")
            popup.dismiss()
            self._load_data()
        except:
            info_popup("Error", "Valores invalidos")

    def _del_avance_hist(self, aid):
        confirm_popup("Eliminar", "Eliminar avance?", lambda: [delete_avance(aid), self._load_data()])

    def _edit_nota_hist(self, nid):
        notas = get_field_notes(self.fecha_spinner.text)
        n = next((x for x in notas if x["id"] == nid), None)
        if not n:
            return
        content = BoxLayout(orientation="vertical", spacing=dp(6), padding=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        txt_i = TextInput(text=n["nota"], height=dp(100), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        content.add_widget(colored_label("Nota", YELLOW, size=12))
        content.add_widget(txt_i)
        btn = cat_button("GUARDAR", lambda x: [update_field_note(nid, txt_i.text.strip()),
                                                popup.dismiss(), self._load_data()])
        content.add_widget(btn)
        popup = Popup(title="Editar nota", content=content, size_hint=[0.9, None], height=dp(280))
        popup.open()

    def _del_nota_hist(self, nid):
        confirm_popup("Eliminar", "Eliminar nota?", lambda: [delete_field_note(nid), self._load_data()])

    def _edit_foto_hist(self, pid):
        fotos = get_photos(self.fecha_spinner.text)
        f = next((x for x in fotos if x["id"] == pid), None)
        if not f:
            return
        content = BoxLayout(orientation="vertical", spacing=dp(6), padding=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        desc_i = TextInput(text=f["descripcion"] or "", height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        content.add_widget(colored_label("Descripcion", YELLOW, size=12))
        content.add_widget(desc_i)
        btn = cat_button("GUARDAR", lambda x: [update_photo(pid, f["sector"], desc_i.text.strip()),
                                                popup.dismiss(), self._load_data()])
        content.add_widget(btn)
        popup = Popup(title="Editar foto", content=content, size_hint=[0.9, None], height=dp(200))
        popup.open()

    def _del_foto_hist(self, pid):
        confirm_popup("Eliminar", "Eliminar foto?", lambda: [delete_photo(pid), self._load_data()])


class AvanceGralScreen(BaseScreen):
    def on_pre_enter(self):
        self.clear_widgets()
        root = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(12), dp(10)])
        root.bind(minimum_height=root.setter("height"))
        self.add_header(root)
        root.add_widget(section_title("AVANCE GENERAL DE PARTIDAS"))

        partidas = get_partidas()
        if not partidas:
            root.add_widget(colored_label("No hay partidas", YELLOW, size=12))
            root.add_widget(back_btn(self.app))
            scroll = ScrollView(size_hint=(1, 1))
            scroll.add_widget(root)
            self.add_widget(scroll)
            return

        fechas = get_registros_fechas()
        for p in partidas:
            total_ejec = 0.0
            total_hrs = 0.0
            for f in fechas:
                reg = get_registro_diario(f)
                if reg:
                    for a in get_avances(reg["id"]):
                        if a["partida_id"] == p["id"]:
                            total_ejec += a["cantidad_ejecutada"]
                            total_hrs += a["horas_trabajadas"]
            pct = (total_ejec / p["metrado_total"] * 100) if p["metrado_total"] > 0 else 0
            txt = (f"{p['codigo']+' - ' if p['codigo'] else ''}{p['nombre']}\n"
                   f"Avance: {total_ejec:.2f} / {p['metrado_total']:.2f} {p['unidad'] or ''} ({pct:.1f}%)")
            root.add_widget(colored_label(txt, WHITE, size=12))
            pb = ProgressBar(max=100, value=min(pct, 100), size_hint_y=None, height=dp(12))
            root.add_widget(pb)

        root.add_widget(Label(size_hint_y=None, height=dp(10)))
        root.add_widget(back_btn(self.app))
        scroll = ScrollView(size_hint=(1, 1))
        scroll.add_widget(root)
        self.add_widget(scroll)


class AdminScreen(BaseScreen):
    def on_pre_enter(self):
        self.clear_widgets()
        root = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(12), dp(10)])
        root.bind(minimum_height=root.setter("height"))
        self.add_header(root)
        root.add_widget(section_title("ADMINISTRAR CATALOGOS"))

        tabs = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(4))
        tab_buttons = [
            ("SECTORES", "sectores"),
            ("MATERIALES", "materiales"),
            ("PARTIDAS", "partidas"),
            ("SUBCONTRATAS", "subcontratas"),
            ("CONFIG", "config"),
        ]
        for txt, key in tab_buttons:
            btn = Button(text=txt, background_color=[0.2, 0.2, 0.2, 1], color=[1, 1, 1, 1],
                         font_size=sp(10), bold=True)
            btn.bind(on_press=lambda x, k=key: self._switch_tab(k))
            tabs.add_widget(btn)
        root.add_widget(tabs)

        self.tab_content = BoxLayout(orientation="vertical", spacing=dp(6), size_hint_y=None)
        self.tab_content.bind(minimum_height=self.tab_content.setter("height"))
        root.add_widget(self.tab_content)
        root.add_widget(Label(size_hint_y=None, height=dp(10)))
        root.add_widget(back_btn(self.app))
        scroll = ScrollView(size_hint=(1, 1))
        scroll.add_widget(root)
        self.add_widget(scroll)
        self._current_tab = "sectores"
        self._show_sectores()

    def _switch_tab(self, tab):
        self._current_tab = tab
        self.tab_content.clear_widgets()
        if tab == "sectores":
            self._show_sectores()
        elif tab == "materiales":
            self._show_materiales()
        elif tab == "partidas":
            self._show_partidas()
        elif tab == "subcontratas":
            self._show_subcontratas()
        elif tab == "config":
            self._show_config()

    def _show_sectores(self):
        tc = self.tab_content
        tc.add_widget(section_title("SECTORES"))
        self.sec_input = TextInput(text="", size_hint_y=None, height=dp(40),
                                    foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        tc.add_widget(colored_label("Nuevo sector"))
        tc.add_widget(self.sec_input)
        tc.add_widget(cat_button("AGREGAR SECTOR", self._add_sector, height=dp(40)))

        for s in get_sectores():
            lbl = colored_label(s["nombre"], WHITE, size=12)
            tc.add_widget(lbl)
            btn_box = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(8))
            btn_box.add_widget(small_btn("EDIT", lambda x, sid=s["id"], nm=s["nombre"]: self._edit_sector(sid, nm)))
            btn_box.add_widget(del_btn(lambda x, sid=s["id"]: self._del_sector(sid)))
            tc.add_widget(btn_box)

    def _add_sector(self, *args):
        if self.sec_input.text.strip():
            save_sector(self.sec_input.text.strip())
            info_popup("OK", "Sector agregado")
            self.sec_input.text = ""
            self._switch_tab("sectores")

    def _edit_sector(self, sid, current_name):
        content = BoxLayout(orientation="vertical", spacing=dp(6), padding=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        nm_i = TextInput(text=current_name, height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        content.add_widget(colored_label("Nombre", YELLOW, size=12))
        content.add_widget(nm_i)
        btn = cat_button("GUARDAR", lambda x: [update_sector(sid, nm_i.text.strip()),
                                                popup.dismiss(), self._switch_tab("sectores")])
        content.add_widget(btn)
        popup = Popup(title="Editar sector", content=content, size_hint=[0.9, None], height=dp(200))
        popup.open()

    def _del_sector(self, sid):
        confirm_popup("Eliminar", "Eliminar sector?", lambda: [delete_sector(sid), self._switch_tab("sectores")])

    def _show_materiales(self):
        tc = self.tab_content
        tc.add_widget(section_title("CATALOGO DE MATERIALES"))
        tc.add_widget(colored_label("Nuevo material", WHITE, size=12))
        grid = GridLayout(cols=2, spacing=dp(4), size_hint_y=None, height=dp(80))
        self.mat_nom = TextInput(text="", height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        self.mat_uni = TextInput(text="", height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        grid.add_widget(colored_label("Nombre", WHITE, size=11))
        grid.add_widget(colored_label("Unidad", WHITE, size=11))
        grid.add_widget(self.mat_nom)
        grid.add_widget(self.mat_uni)
        tc.add_widget(grid)
        tc.add_widget(cat_button("AGREGAR MATERIAL", self._add_mat_cat, height=dp(40)))
        tc.add_widget(cat_button("IMPORTAR DESDE EXCEL", self._import_mat_excel, height=dp(36)))

        for m in get_materiales_catalogo():
            lbl = colored_label(f"{m['nombre']} - {m['unidad']}", WHITE, size=12)
            tc.add_widget(lbl)
            btn_box = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(8))
            btn_box.add_widget(small_btn("EDIT", lambda x, mid=m["id"], nm=m["nombre"], un=m["unidad"]: self._edit_mat_cat(mid, nm, un)))
            btn_box.add_widget(del_btn(lambda x, mid=m["id"]: self._del_mat_cat(mid)))
            tc.add_widget(btn_box)

    def _add_mat_cat(self, *args):
        if self.mat_nom.text.strip() and self.mat_uni.text.strip():
            save_material_catalogo(self.mat_nom.text.strip(), self.mat_uni.text.strip())
            info_popup("OK", "Material agregado")
            self.mat_nom.text = ""
            self.mat_uni.text = ""
            self._switch_tab("materiales")

    def _edit_mat_cat(self, mid, current_name, current_unidad):
        content = BoxLayout(orientation="vertical", spacing=dp(6), padding=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        nm_i = TextInput(text=current_name, height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        un_i = TextInput(text=current_unidad, height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        content.add_widget(colored_label("Nombre", YELLOW, size=12))
        content.add_widget(nm_i)
        content.add_widget(colored_label("Unidad", YELLOW, size=12))
        content.add_widget(un_i)
        btn = cat_button("GUARDAR", lambda x: [update_material_catalogo(mid, nm_i.text.strip(), un_i.text.strip()),
                                                popup.dismiss(), self._switch_tab("materiales")])
        content.add_widget(btn)
        popup = Popup(title="Editar material", content=content, size_hint=[0.9, None], height=dp(300))
        popup.open()

    def _del_mat_cat(self, mid):
        confirm_popup("Eliminar", "Eliminar material del catalogo?",
                      lambda: [delete_material_catalogo(mid), self._switch_tab("materiales")])

    def _import_mat_excel(self, *args):
        content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        content.add_widget(colored_label("Selecciona archivo Excel", YELLOW, size=14))
        content.add_widget(colored_label("Columnas: Nombre, Unidad", WHITE, size=12))
        filechooser = FileChooserListView(path=os.path.expanduser("~"), size_hint_y=None, height=dp(300),
                                          filters=["*.xlsx", "*.xls"])
        content.add_widget(filechooser)
        btn_importar = cat_button("IMPORTAR", lambda x: self._do_import_mat(filechooser.path, filechooser.selection, popup))
        content.add_widget(btn_importar)
        popup = Popup(title="Importar materiales", content=content, size_hint=[0.95, 0.8])
        popup.open()

    def _do_import_mat(self, folder, selection, popup):
        if not selection:
            info_popup("Error", "Selecciona un archivo")
            return
        filepath = selection[0]
        if not filepath.endswith((".xlsx", ".xls")):
            info_popup("Error", "Debe ser archivo .xlsx")
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip().upper() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if "NOMBRE" not in headers or "UNIDAD" not in headers:
                info_popup("Error", "El Excel debe tener columnas: Nombre, Unidad")
                wb.close()
                return
            col_nom = headers.index("NOMBRE")
            col_uni = headers.index("UNIDAD")
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                nombre = str(row[col_nom] or "").strip()
                unidad = str(row[col_uni] or "").strip()
                if nombre and unidad:
                    save_material_catalogo(nombre, unidad)
                    count += 1
            wb.close()
            popup.dismiss()
            info_popup("OK", f"{count} materiales importados")
            self._switch_tab("materiales")
        except Exception as e:
            info_popup("Error", f"Error al importar: {e}")

    def _import_par_excel(self, *args):
        content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        content.add_widget(colored_label("Selecciona archivo Excel", YELLOW, size=14))
        content.add_widget(colored_label("Columnas: Codigo, Nombre, Descripcion, Unidad, Metrado", WHITE, size=12))
        filechooser = FileChooserListView(path=os.path.expanduser("~"), size_hint_y=None, height=dp(300),
                                          filters=["*.xlsx", "*.xls"])
        content.add_widget(filechooser)
        btn_importar = cat_button("IMPORTAR", lambda x: self._do_import_par(filechooser.path, filechooser.selection, popup))
        content.add_widget(btn_importar)
        popup = Popup(title="Importar partidas", content=content, size_hint=[0.95, 0.8])
        popup.open()

    def _do_import_par(self, folder, selection, popup):
        if not selection:
            info_popup("Error", "Selecciona un archivo")
            return
        filepath = selection[0]
        if not filepath.endswith((".xlsx", ".xls")):
            info_popup("Error", "Debe ser archivo .xlsx")
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip().upper() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if "NOMBRE" not in headers:
                info_popup("Error", "El Excel debe tener columna: Nombre")
                wb.close()
                return
            col_idx = {h: i for i, h in enumerate(headers)}
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                nombre = str(row[col_idx["NOMBRE"]] or "").strip() if "NOMBRE" in col_idx and col_idx["NOMBRE"] < len(row) else ""
                if not nombre:
                    continue
                cod = str(row[col_idx["CODIGO"]] or "").strip() if "CODIGO" in col_idx and col_idx["CODIGO"] < len(row) else ""
                desc = str(row[col_idx["DESCRIPCION"]] or "").strip() if "DESCRIPCION" in col_idx and col_idx["DESCRIPCION"] < len(row) else ""
                uni = str(row[col_idx["UNIDAD"]] or "").strip() if "UNIDAD" in col_idx and col_idx["UNIDAD"] < len(row) else ""
                met = float(row[col_idx["METRADO"]]) if "METRADO" in col_idx and col_idx["METRADO"] < len(row) and row[col_idx["METRADO"]] is not None else 0.0
                save_partida(cod, nombre, desc, None, uni, met)
                count += 1
            wb.close()
            popup.dismiss()
            info_popup("OK", f"{count} partidas importados")
            self._switch_tab("partidas")
        except Exception as e:
            info_popup("Error", f"Error al importar: {e}")

    def _show_partidas(self):
        tc = self.tab_content
        tc.add_widget(section_title("PARTIDAS"))
        frm = GridLayout(cols=2, spacing=dp(4), size_hint_y=None, height=dp(240))
        self.par_cod = TextInput(text="", height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        self.par_nom = TextInput(text="", height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        self.par_desc = TextInput(text="", height=dp(60), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        self.par_uni = TextInput(text="", height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        self.par_met = TextInput(text="0.0", height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        labels = ["Codigo", "Nombre", "Descripcion", "Unidad", "Metrado"]
        inputs = [self.par_cod, self.par_nom, self.par_desc, self.par_uni, self.par_met]
        for lbl, inp in zip(labels, inputs):
            tc.add_widget(colored_label(lbl, WHITE, size=11))
            tc.add_widget(inp)
        tc.add_widget(cat_button("AGREGAR PARTIDA", self._add_partida, height=dp(40)))
        tc.add_widget(cat_button("IMPORTAR DESDE EXCEL", self._import_par_excel, height=dp(36)))

        for p in get_partidas():
            txt = f"{p['codigo']+' - ' if p['codigo'] else ''}{p['nombre']}\n{p['descripcion'] or ''}\nUnidad: {p['unidad'] or 'N/A'} | Metrado: {p['metrado_total']}"
            tc.add_widget(colored_label(txt, WHITE, size=11))
            btn_box = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(8))
            btn_box.add_widget(small_btn("EDIT", lambda x, pid=p["id"]: self._edit_partida(pid)))
            btn_box.add_widget(del_btn(lambda x, pid=p["id"]: self._del_partida(pid)))
            tc.add_widget(btn_box)

    def _add_partida(self, *args):
        try:
            met = float(self.par_met.text)
        except:
            met = 0.0
        save_partida(self.par_cod.text, self.par_nom.text, self.par_desc.text, None, self.par_uni.text, met)
        info_popup("OK", "Partida agregada")
        self._switch_tab("partidas")

    def _edit_partida(self, pid):
        partidas = get_partidas()
        p = next((x for x in partidas if x["id"] == pid), None)
        if not p:
            return
        content = BoxLayout(orientation="vertical", spacing=dp(6), padding=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        cod_i = TextInput(text=p["codigo"] or "", height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        nom_i = TextInput(text=p["nombre"], height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        desc_i = TextInput(text=p["descripcion"] or "", height=dp(60), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        uni_i = TextInput(text=p["unidad"] or "", height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        met_i = TextInput(text=str(p["metrado_total"]), height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        for lbl, w in [("Codigo", cod_i), ("Nombre", nom_i), ("Descripcion", desc_i),
                        ("Unidad", uni_i), ("Metrado", met_i)]:
            content.add_widget(colored_label(lbl, YELLOW, size=12))
            content.add_widget(w)
        btn = cat_button("GUARDAR", lambda x: [update_partida(pid, cod_i.text, nom_i.text, desc_i.text, None,
                                                               uni_i.text, float(met_i.text or "0")),
                                                popup.dismiss(), self._switch_tab("partidas")])
        content.add_widget(btn)
        popup = Popup(title="Editar partida", content=content, size_hint=[0.9, None], height=dp(500))
        popup.open()

    def _del_partida(self, pid):
        confirm_popup("Eliminar", "Eliminar partida?", lambda: [delete_partida(pid), self._switch_tab("partidas")])

    def _show_subcontratas(self):
        tc = self.tab_content
        tc.add_widget(section_title("SUBCONTRATAS"))
        self.subc_nom = TextInput(text="", height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        self.subc_resp = TextInput(text="", height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        self.subc_tel = TextInput(text="", height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        for lbl, w in [("Nombre", self.subc_nom), ("Responsable", self.subc_resp), ("Telefono", self.subc_tel)]:
            tc.add_widget(colored_label(lbl, WHITE, size=11))
            tc.add_widget(w)
        tc.add_widget(cat_button("AGREGAR SUBCONTRATA", self._add_subc, height=dp(40)))
        for s in get_subcontratas():
            txt = f"{s['nombre']}\nResp: {s['responsable'] or 'N/A'} | Tel: {s['telefono'] or 'N/A'}"
            tc.add_widget(colored_label(txt, WHITE, size=11))
            btn_box = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(8))
            btn_box.add_widget(small_btn("EDIT", lambda x, sid=s["id"]: self._edit_subc(sid)))
            btn_box.add_widget(del_btn(lambda x, sid=s["id"]: self._del_subc(sid)))
            tc.add_widget(btn_box)

    def _add_subc(self, *args):
        if self.subc_nom.text.strip():
            save_subcontrata(self.subc_nom.text.strip(), self.subc_resp.text, self.subc_tel.text)
            info_popup("OK", "Subcontrata agregada")
            self._switch_tab("subcontratas")

    def _edit_subc(self, sid):
        subcs = get_subcontratas()
        s = next((x for x in subcs if x["id"] == sid), None)
        if not s:
            return
        content = BoxLayout(orientation="vertical", spacing=dp(6), padding=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        nm_i = TextInput(text=s["nombre"], height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        resp_i = TextInput(text=s["responsable"] or "", height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        tel_i = TextInput(text=s["telefono"] or "", height=dp(36), foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        for lbl, w in [("Nombre", nm_i), ("Responsable", resp_i), ("Telefono", tel_i)]:
            content.add_widget(colored_label(lbl, YELLOW, size=12))
            content.add_widget(w)
        btn = cat_button("GUARDAR", lambda x: [update_subcontrata(sid, nm_i.text.strip(), resp_i.text, tel_i.text),
                                                popup.dismiss(), self._switch_tab("subcontratas")])
        content.add_widget(btn)
        popup = Popup(title="Editar subcontrata", content=content, size_hint=[0.9, None], height=dp(350))
        popup.open()

    def _del_subc(self, sid):
        confirm_popup("Eliminar", "Eliminar subcontrata?",
                      lambda: [delete_subcontrata(sid), self._switch_tab("subcontratas")])

    def _show_config(self):
        tc = self.tab_content
        tc.add_widget(section_title("CONFIGURACION"))
        config = get_all_config()

        self.cfg_resp_nom = TextInput(text=config.get("responsable_nombre", ""), height=dp(36),
                                       foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        self.cfg_resp_cargo = TextInput(text=config.get("responsable_cargo", ""), height=dp(36),
                                         foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        self.cfg_res_nom = TextInput(text=config.get("residente_nombre", ""), height=dp(36),
                                      foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        self.cfg_res_cargo = TextInput(text=config.get("residente_cargo", ""), height=dp(36),
                                        foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        self.cfg_proy = TextInput(text=config.get("proyecto_nombre", ""), height=dp(60),
                                   foreground_color=[1,1,1,1], background_color=[0,0,0,1])
        self.cfg_cui = TextInput(text=config.get("cui", ""), height=dp(36),
                                  foreground_color=[1,1,1,1], background_color=[0,0,0,1])

        tc.add_widget(colored_label("RESPONSABLE (DE)", YELLOW, bold=True, size=13))
        tc.add_widget(colored_label("Nombre"))
        tc.add_widget(self.cfg_resp_nom)
        tc.add_widget(colored_label("Cargo"))
        tc.add_widget(self.cfg_resp_cargo)

        tc.add_widget(colored_label("RESIDENTE (A)", YELLOW, bold=True, size=13))
        tc.add_widget(colored_label("Nombre"))
        tc.add_widget(self.cfg_res_nom)
        tc.add_widget(colored_label("Cargo"))
        tc.add_widget(self.cfg_res_cargo)

        tc.add_widget(colored_label("PROYECTO", YELLOW, bold=True, size=13))
        tc.add_widget(colored_label("Nombre"))
        tc.add_widget(self.cfg_proy)
        tc.add_widget(colored_label("CUI"))
        tc.add_widget(self.cfg_cui)
        tc.add_widget(cat_button("GUARDAR CONFIGURACION", self._save_config, height=dp(40)))

    def _save_config(self, *args):
        set_config("responsable_nombre", self.cfg_resp_nom.text)
        set_config("responsable_cargo", self.cfg_resp_cargo.text)
        set_config("residente_nombre", self.cfg_res_nom.text)
        set_config("residente_cargo", self.cfg_res_cargo.text)
        set_config("proyecto_nombre", self.cfg_proy.text)
        set_config("cui", self.cfg_cui.text)
        info_popup("OK", "Configuracion guardada")


class ControlObraApp(App):
    def build(self):
        self.title = "CONTROL DE OBRA"
        if platform == "android":
            Window.softinput_mode = "below_target"

        app_dir = self.user_data_dir
        set_db_path(os.path.join(app_dir, "db_obra.sqlite"))
        init_db()

        sm = ScreenManager(transition=SlideTransition(direction="left"))
        sm.add_widget(MenuScreen(name="menu"))
        sm.add_widget(RegistroScreen(name="registro"))
        sm.add_widget(MaterialesScreen(name="materiales"))
        sm.add_widget(NotasScreen(name="notas"))
        sm.add_widget(FotosScreen(name="fotos"))
        sm.add_widget(InformeScreen(name="informe"))
        sm.add_widget(InformeSemanalScreen(name="informe_semanal"))
        sm.add_widget(HistorialScreen(name="historial"))
        sm.add_widget(AdminScreen(name="admin"))
        sm.add_widget(AvanceGralScreen(name="avance_gral"))
        return sm

    def switch_screen(self, screen_name):
        self.root.current = screen_name


if __name__ == "__main__":
    ControlObraApp().run()
